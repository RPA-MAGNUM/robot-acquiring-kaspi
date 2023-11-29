import sys


def split_branches(src_file, dst_dir):
    print('!!!', src_file)
    print('!!!', dst_dir)
    from pathlib import Path
    from openpyxl.reader.excel import load_workbook
    from openpyxl.workbook import Workbook

    src_file = Path(src_file)
    dst_dir = Path(dst_dir)
    wb = load_workbook(src_file.__str__())
    ws = wb.active
    vs = list(ws.values)
    start_row = [n for n, r in enumerate(vs) if r[0] == '#'][0]

    branches = dict()
    for n, row in enumerate(vs):
        if n <= start_row:
            continue
        if not branches.get(row[1]):
            branches[row[1]] = [row]
        else:
            branches[row[1]].append(row)

    for n, branch in enumerate(list(branches.keys())):
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = ws.title
        [new_ws.append(vs[n]) for n in range(start_row + 1)]
        [new_ws.append(r) for r in branches[branch]]
        new_wb.save(dst_dir.joinpath(f'{src_file.stem}_{n}{src_file.suffix}'))
        new_wb.close()
    wb.close()
    del wb
    del ws
    del vs
    del branches


if __name__ == '__main__':
    if len(sys.argv) == 3:
        split_branches(sys.argv[1], sys.argv[2])
