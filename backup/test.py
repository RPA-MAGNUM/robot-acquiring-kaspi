from pathlib import Path

from openpyxl import load_workbook

a = 1
for file in Path(r"C:\Users\rpa.robot\PycharmProjects\robot-acquiring-kaspi\manual\done").glob('*.xlsx'):
    print(a)
    a+= 1
    wb = load_workbook(file.__str__())
    ws = wb.active
    cell = ws.cell(8, 2).value
    if 'tr-t Kuldzhinskiy, uch 106' in cell:
        print(file)
    wb.close()