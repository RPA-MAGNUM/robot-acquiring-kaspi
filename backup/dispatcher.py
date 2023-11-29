import datetime
import os
import uuid
from pathlib import Path

import openpyxl
import pandas as pd
import psycopg2

from config import logger, db_host, db_port, db_name, db_user, db_pass, months, \
    main_directory_folder, robot_name, str_date_working_file, str_parking_folder


def table_create():
    """
        Just simple table creation if not exists
        """
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''CREATE TABLE IF NOT EXISTS ROBOT.{robot_name.replace("-", "_")} (
        id text PRIMARY KEY,
        status text,
        retry_count INTEGER,
        error_message text,
        comments text,
        execution_time text,
        finish_date text,
        date_created text,
        executor_name text,
        document_number text,
        debit_credit_sum text,
        purpose_of_payment text, 
        operation_date text,
        folder_date text) '''

    c = conn.cursor()
    c.execute(table_create_query)
    conn.commit()
    table_create_query = f'''CREATE TABLE IF NOT EXISTS ROBOT.{robot_name.replace("-", "_")}_parking (
            id text PRIMARY KEY,
            status text,
            retry_count INTEGER,
            error_message text,
            comments text,
            execution_time text,
            finish_date text,
            date_created text,
            executor_name text,
            operation_date text,
            folder_date text,
            file_path text) '''
    # * Parking table
    c.execute(table_create_query)
    conn.commit()
    c.close()
    conn.close()


def search_files(root_dir, keyword):
    matching_files = []
    for dirpath, dirnames, filenames in os.walk(root_dir):
        for filename in filenames:
            if keyword in filename:
                matching_files.append(os.path.join(dirpath, filename))
    return matching_files


def dispatch():
    logger.info("Starting dispatcher")

    # logger.info("0) START table_create - простое создание таблицы в postgre")
    table_create()
    # logger.info("0) END table_create - простое создание таблицы в postgre")

    # logger.info(f"1) START чтение маппинг файла для выписки дат {str_date_working_file}")
    # * Read mapping file for date determination.
    print(f"{str_date_working_file}")
    wb = openpyxl.load_workbook(str_date_working_file, data_only=True)

    ws = wb['Каспи']
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    c = conn.cursor()
    holiday = True
    operation_date = None
    folder_dates = None

    today = datetime.datetime.now().date()
    # today = datetime.datetime.strptime("12.07.2023", "%d.%m.%Y").date()
    today_str = datetime.datetime.now().strftime("%d.%m.%Y")
    # today_str = "12.07.2023"
    for idx, row in enumerate(ws.iter_rows(min_row=1)):

        if isinstance(row[0].value, datetime.datetime):
            if row[0].value.date() == today:
                print(f"datetime {row[0].value}")
                if row[2].value == "выходной":
                    logger.info("Сегодня выходной")
                    return
                else:
                    holiday = False
                    logger.info("Сегодня рабочий день")

        else:
            print(f"str {row[0].value}")
            if row[0].value == today_str:
                if row[2].value == "выходной":
                    logger.info("Сегодня выходной")
                    return
                else:
                    holiday = False
                    logger.info("Сегодня рабочий день")

            else:
                continue
        if not holiday:
            operation_date = row[1].value
            folder_dates = row[4].value
            logger.info(f"Operation date: {operation_date}")
            break
    if holiday:
        logger.info("Не нашли дату")
        return
    if isinstance(operation_date, datetime.datetime):
        folder_date = operation_date.strftime("%d.%m.%Y")
        operation_date = operation_date.strftime("%d.%m.%Y")
    else:
        folder_date = operation_date

    # я думал что во вт нужно будет отрабатывать несколько файлов, но оказалось что файл один и называется он комплексно
    # written by ; if many

    if isinstance(folder_dates, datetime.datetime):
        folder_date = folder_dates.strftime("%d.%m.%Y")
        folder_dates = folder_date
    else:
        if ";" in folder_dates:
            folder_date = str(folder_dates).split(";")[-1].strip()

    datetime_obj = datetime.datetime.strptime(folder_date, "%d.%m.%Y")
    current_month: int = datetime_obj.month
    current_month_folder_name: str = months[current_month]
    current_year: int = datetime_obj.year
    # folder_day = datetime_obj.strftime("%d.%m")

    folder_path = Path(main_directory_folder).joinpath(str(current_year),
                                                       f"{current_month_folder_name} {current_year}")
    files = os.listdir(str(folder_path))
    file_found = False
    file_path = ''
    for item in files:
        if operation_date in item:
            file_found = True
            file_path = folder_path.joinpath(item)
            break
    if not file_found:
        logger.info(f"Каспи Выписка не найдена за день {folder_date}")
        # TODO raise Exception
        return
    # logger.info(f"1) END чтение маппинг файла для выписки дат {str_date_working_file}")

    # logger.info(f"2) START чтение мэйн файла, если нет такого создаст сам {file_path}")
    # *  Find the corresponding main file, and if not exists, create new from mapping file.
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    transaction_count = 0
    for idx, row in enumerate(ws.iter_rows(min_row=14)):
        # TODO to do it I need access to the shared folder
        if row[0].value:
            document_number = str(row[0].value)
            debit = str(row[2].value).strip() if row[2].value else None
            credit = str(row[3].value).strip() if row[3].value else None
            debit_credit_sum = debit if debit else credit

            purpose_of_payment = str(row[8].value).strip()
            print(f"Сумма: {debit_credit_sum} {purpose_of_payment}")

            # * Need to create transactions based on mapping data
            # * Establish connection

            str_now = datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')

            # * need to check whether we already have a trans in db
            find_query = f"Select id from ROBOT.{robot_name.replace('-', '_')} where " \
                         f"operation_date='{operation_date}' AND document_number='{document_number}'"

            c.execute(find_query)
            result = c.fetchone()
            if result is None:
                logger.info(f"operation date at the end{operation_date}")

                # * insert a transaction to db
                insert_query = f"""Insert into ROBOT.{robot_name.replace('-', '_')} (id, document_number, debit_credit_sum, purpose_of_payment, folder_date, operation_date,  status, retry_count, date_created) values ('{uuid.uuid4()}', '{document_number}', '{debit_credit_sum}', '{purpose_of_payment}', '{folder_dates}', '{operation_date}' , 'New', 0, '{str_now}') """
                c.execute(insert_query)
                conn.commit()
                transaction_count += 1

            # * close db connection

    logger.info(f"Added {transaction_count} rows to  the DB")
    # logger.info(f"2) END чтение мэйн файла, если нет такого создаст сам {file_path}")

    # * Добавляем в бд парковки
    wb = openpyxl.load_workbook(str_date_working_file, data_only=True)

    # Step 1: Определить рабочий день-----------------
    ws = wb['Каспи']
    operation_date = None
    today = datetime.datetime.now().strftime("%d.%m.%Y")
    # today = "02.10.2023"

    for idx, row in enumerate(ws.iter_rows(min_row=0)):
        date_m = row[0].value
        if isinstance(date_m, datetime.datetime):
            date_m = date_m.strftime("%d.%m.%Y")

        if date_m == today:
            if row[2].value == "выходной":
                logger.info("Сегодня выходной")
                return
            else:
                logger.info("Сегодня рабочий день")
                # holiday = False

        else:
            continue
        operation_date = row[1].value
        if isinstance(operation_date, datetime.datetime):
            operation_date = operation_date.strftime("%d.%m.%Y")

        print(f"Operation_date: {operation_date}")

        folder_dates = row[3].value  # written by ; if many
        if isinstance(folder_dates, datetime.datetime):
            folder_dates = [folder_dates.strftime("%d.%m.%Y")]
        else:
            folder_dates = str(folder_dates).split(";")

    transaction_count = 0
    for folder_date in folder_dates:
        # step 3: Далее отрабатываем файлы парковки по очередности по филиально
        search_date = str(folder_date).strip()
        str_now = datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        datetime_obj = datetime.datetime.strptime(search_date, "%d.%m.%Y")
        current_year = datetime_obj.year
        current_month: int = datetime_obj.month
        current_month_folder_name: str = months[current_month]
        root_dir = Path(str_parking_folder).joinpath(str(current_year), current_month_folder_name)
        list_of_parking_files = search_files(str(root_dir), search_date)
        for each in list_of_parking_files:
            df = pd.read_excel(Path(each).__str__())
            vs = list(df.values.tolist())

            status__ = 'Fail' if (vs[1][0] == 'Итого:' and not vs[1][7]) else 'New'
            # * need to check whether we already have a trans in db
            find_query = f"Select id from ROBOT.{robot_name.replace('-', '_')}_parking where operation_date='{operation_date}' AND folder_date='{folder_date}' AND file_path='{each}'"

            c.execute(find_query)
            result = c.fetchone()
            if result is None:
                logger.info(f"operation date at the end{operation_date}")

                # * insert a transaction to db
                insert_query = f"""Insert into ROBOT.{robot_name.replace('-', '_')}_parking (id,  folder_date, operation_date,  status, retry_count, date_created, file_path) values ('{uuid.uuid4()}',
                               '{search_date}', '{operation_date}' , '{status__}', 0, '{str_now}', '{each}') """
                print("INset query parking ", insert_query)
                c.execute(insert_query)
                conn.commit()
                transaction_count += 1
    logger.info(f"added {transaction_count} parkings to db")
    logger.info("Dispatcher ended")
    c.close()
    conn.close()
    # logger.info("Закончили")


if __name__ == '__main__':
    dispatch()
