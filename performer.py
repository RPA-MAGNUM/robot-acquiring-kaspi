import datetime
import os
import shutil
import time
import traceback
from pathlib import Path
# from subprocess import Popen

import openpyxl
import pandas as pd
import psycopg2
import pyautogui
from dateutil.relativedelta import relativedelta

from config import logger, db_host, db_port, db_name, db_user, db_pass, ip_address, robot_name, smtp_author, \
    smtp_host, to_whom, cc_whom, robot_name_russian, str_date_working_file, temp_folder, str_path_mapping_excel_file, \
    upload_folder, months, str_sales_folder, \
    screenshots_folder, transaction_retry_count, owa_username
from core import Odines, Cursor
from rpamini import send_message_by_smtp, try_except_decorator, clipboard_set, clipboard_get, BusinessException, \
    check_file_downloaded, fix_excel_file_error, retry_n_times, kill_process_list


class Transaction:
    def __init__(self, row):
        # * template columns in db
        self.id = row[0]
        # logger.warning(f'ID ---> {self.id}')
        self.status = row[1]
        self.retry_count = row[2]
        self.error_message = row[3]
        self.comments = row[4]
        self.execution_time = row[5]
        self.finish_date = row[6]
        self.date_created = row[7]
        self.executor_name = row[8]
        self.start_time = None
        # * ----------------------
        self.operation_date = row[12]
        self.folder_date = row[13]
        self.needed_sum = row[10]
        self.purpose_of_payment = row[11]
        # logger.warning(f'TYPE ---> {self.purpose_of_payment}')
        print(f"self.operation_date: {self.operation_date}")
        print(f'self.folder_date: {self.folder_date}')
        print(f'needed_sum: {self.needed_sum}')
        print(f'purpose_of_payment: {self.purpose_of_payment}')

    def write_off_1cc(self, list_of_demand):
        """Example function"""

        # * selectors
        # opened_table_selector = {"title": "", "class_name": "", "control_type": "Table", "visible_only": True,
        #                          "enabled_only": True, "found_index": 0}
        filter_selector = {"title": "Установить отбор и сортировку списка...", "class_name": "",
                           "control_type": "Button",
                           "visible_only": True, "enabled_only": True, "found_index": 0}
        filter_whole_wnd_selector = {"title": "Отбор и сортировка", "class_name": "V8NewLocalFrameBaseWnd",
                                     "control_type": "Window", "visible_only": True, "enabled_only": True,
                                     "found_index": 0}

        filter_wnd_counterparty_field_selector = {"title": "", "class_name": "", "control_type": "Edit",
                                                  "visible_only": True, "enabled_only": True, "found_index": 15}
        filter_wnd_operation_type_selector = {"title": "", "class_name": "", "control_type": "Edit",
                                              "visible_only": True,
                                              "enabled_only": True, "found_index": 11}
        # filter_wnd_comments_selector = {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
        #                                 "enabled_only": True, "found_index": 13}

        filter_ok_button_selector = {"title": "OK", "class_name": "", "control_type": "Button", "visible_only": True,
                                     "enabled_only": True, "found_index": 0}
        first_entry = {"title_re": "^.* Дата", "class_name": "", "control_type": "Custom", "visible_only": True,
                       "enabled_only": True, "found_index": 0}

        inner_selector = {"title": "Основная", "class_name": "", "control_type": "Tab", "visible_only": True,
                          "enabled_only": True, "found_index": 0}

        date_1 = 0

        sum_field = 0
        comment_f = 0

        if list_of_demand[0] == "Прочее списание безналичных денежных средств":
            date_1 = 10

            sum_field = 7
            comment_f = 11
        elif list_of_demand[0] == "Возврат платежа по эквайринговым операциям POS":
            print("Возврат платежа по эквайринговым операциям POS")
            date_1 = 9

            sum_field = 6
            comment_f = 10

        else:
            logger.info("Ошибка в списание!!!!!")
            # logger.info(f"list of demand0[0] {list_of_demand[0]}")
        date_1_field_selector = {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
                                 "enabled_only": True, "found_index": date_1}

        sum_field_selector = {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
                              "enabled_only": True, "found_index": sum_field}
        comment_field_selector = {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
                                  "enabled_only": True, "found_index": comment_f}

        ok_button_selector = {"title": "ОК", "class_name": "", "control_type": "Button", "visible_only": True,
                              "enabled_only": True, "found_index": 0}

        app = Odines()
        app.run()

        app.navigate("Банк и касса", "Платежный ордер списание денежных средств", maximize_innder=True)

        # table_element = app.find_element(opened_table_selector)

        app.find_element(filter_selector).click()
        # todo switch parent
        app.parent_switch(filter_whole_wnd_selector, resize=True)

        # app.find_element(filter_wnd_comments_selector).type_keys(purpose_of_payment, protect_first=True, clear=True,
        #                                                          click=True)
        time.sleep(1)
        print(list_of_demand)
        app.find_element(filter_wnd_operation_type_selector).type_keys(list_of_demand[0], app.keys.TAB,
                                                                       protect_first=True, clear=True,
                                                                       click=True)
        app.find_element(filter_wnd_counterparty_field_selector).type_keys(list_of_demand[1], app.keys.TAB,
                                                                           protect_first=True, clear=True,
                                                                           click=True)
        app.find_element(filter_ok_button_selector).click()

        app.parent_back(1)

        app.find_element(first_entry).type_keys(app.keys.END, app.keys.F9, click=True, clear=False, protect_first=False)

        error_notification = app.wait_element(
            {"title": "", "class_name": "", "control_type": "Document", "visible_only": True, "enabled_only": True,
             "found_index": 0, "parent": app.root}, timeout=3)
        if error_notification:
            comment = "Документ только для чтения"
            self.comments = comment
            logger.info(comment)
            app.quit()
            return False
        app.parent_switch(app.find_element(inner_selector).parent(4))

        date_el = app.find_element(date_1_field_selector)
        date_el.draw_outline()
        time.sleep(1)
        date_el.type_keys(self.operation_date, app.keys.TAB, click=True, clear=True)
        print(date_1_field_selector)
        sum_el = app.find_element(sum_field_selector)
        sum_el.draw_outline()
        time.sleep(1)
        sum_el.type_keys(self.needed_sum, app.keys.TAB, clear=True, click=True,
                         protect_first=True)

        comment_el = app.find_element(comment_field_selector)
        comment_el.draw_outline()
        time.sleep(1)
        comment_el.type_keys(self.purpose_of_payment, app.keys.TAB,
                             protect_first=True,
                             click=True, clear=True)

        # TODO turn on later
        app.find_element(ok_button_selector).click()
        # * Конфликт блокировок нужно чекнуть
        try_count = 30
        while True:
            try_count -= 1
            if app.wait_element(ok_button_selector, until=False):
                break
            if app.wait_element(
                    {"title_re": "Конфликт блокировок.*", "class_name": "", "control_type": "Pane",
                     "visible_only": True,
                     "enabled_only": True, "found_index": 0}, timeout=3):
                logger.info("Конфлик блокировок")
                if try_count < 0:
                    self.error_message = "Конфикт блокировок"
                    app.quit()
                    raise BusinessException("Не присвоен номер Конфликт блокировок", "Исходщие")
        app.quit()
        return True

    def get_incoming_1c(self, list_of_demand):
        """Example function"""
        # logger.info("Входящие")
        # * selectors
        opened_table_selector = {"title": "", "class_name": "", "control_type": "Table", "visible_only": True,
                                 "enabled_only": True, "found_index": 0}
        filter_selector = {"title": "Установить отбор и сортировку списка...", "class_name": "",
                           "control_type": "Button",
                           "visible_only": True, "enabled_only": True, "found_index": 0}
        filter_whole_wnd_selector = {"title": "Отбор и сортировка", "class_name": "V8NewLocalFrameBaseWnd",
                                     "control_type": "Window", "visible_only": True, "enabled_only": True,
                                     "found_index": 0}

        filter_wnd_counterparty_field_selector = {"title": "", "class_name": "", "control_type": "Edit",
                                                  "visible_only": True, "enabled_only": True, "found_index": 15}
        filter_wnd_operation_type_selector = {"title": "", "class_name": "", "control_type": "Edit",
                                              "visible_only": True,
                                              "enabled_only": True, "found_index": 9}
        # filter_wnd_comments_selector = {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
        #                                 "enabled_only": True, "found_index": 13}

        filter_ok_button_selector = {"title": "OK", "class_name": "", "control_type": "Button", "visible_only": True,
                                     "enabled_only": True, "found_index": 0}
        first_entry = {"title_re": "^.* Дата", "class_name": "", "control_type": "Custom", "visible_only": True,
                       "enabled_only": True, "found_index": 0}

        inner_selector = {"title": "Основная", "class_name": "", "control_type": "Tab", "visible_only": True,
                          "enabled_only": True, "found_index": 0}
        date_1 = 0
        date_2 = 0
        sum_field = 0
        comment_f = 0

        if list_of_demand[0] == "Поступление оплаты по эквайрингу":
            date_1 = 8
            date_2 = 1
            sum_field = 9
            comment_f = 2
        elif list_of_demand[0] == "Возврат денежных средств поставщиком" or \
                list_of_demand[0] == "Прочее поступление безналичных денежных средств":
            date_1 = 9
            date_2 = 1
            sum_field = 10
            comment_f = 2

        date_1_field_selector = {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
                                 "enabled_only": True, "found_index": date_1}
        date_2_field_selector = {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
                                 "enabled_only": True, "found_index": date_2}
        sum_field_selector = {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
                              "enabled_only": True, "found_index": sum_field}
        comment_field_selector = {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
                                  "enabled_only": True, "found_index": comment_f}

        ok_button_selector = {"title": "ОК", "class_name": "", "control_type": "Button", "visible_only": True,
                              "enabled_only": True, "found_index": 0}

        app = Odines()
        app.run()

        app.navigate("Банк и касса", "Платежное поручение входящее", maximize_innder=True)

        app.find_element(opened_table_selector)

        app.find_element(filter_selector).click()
        # todo switch parent
        app.parent_switch(filter_whole_wnd_selector, resize=True)

        # app.find_element(filter_wnd_comments_selector).type_keys(purpose_of_payment, protect_first=True, clear=True,
        #                                                          click=True)
        app.find_element(filter_wnd_operation_type_selector).type_keys(list_of_demand[0], app.keys.TAB,
                                                                       protect_first=True, clear=True,
                                                                       click=True)
        app.find_element(filter_wnd_counterparty_field_selector).type_keys(list_of_demand[1], app.keys.TAB,
                                                                           protect_first=True, clear=True,
                                                                           click=True)
        app.find_element(filter_ok_button_selector).click()

        app.parent_back(1)

        app.find_element(first_entry).type_keys(app.keys.END, app.keys.F9, click=True, clear=False, protect_first=False)
        # wnd_whole = app.parent_switch(table_element.parent(4), maximize=True)

        error_notification = app.wait_element(
            {"title": "", "class_name": "", "control_type": "Document", "visible_only": True, "enabled_only": True,
             "found_index": 0, "parent": app.root}, timeout=3)
        if error_notification:
            comment = "Документ только для чтения"
            self.comments = comment
            logger.info(comment)
            app.quit()
            return False
        app.parent_switch(app.find_element(inner_selector).parent(4))
        # app.maximize_inner()
        app.find_element(date_1_field_selector).type_keys(self.operation_date, app.keys.TAB, click=True, clear=True)

        app.find_element(date_2_field_selector).type_keys(self.operation_date, app.keys.TAB, click=True, clear=True)

        app.find_element(sum_field_selector).type_keys(self.needed_sum, app.keys.TAB, clear=True, click=True,
                                                       protect_first=True)

        app.find_element(comment_field_selector).type_keys(self.purpose_of_payment, app.keys.TAB,
                                                           protect_first=True,
                                                           click=True, clear=True)

        app.find_element(ok_button_selector).click()
        # * Конфликт блокировок нужно чекнуть
        try_count = 30
        while True:
            try_count -= 1
            if app.wait_element(ok_button_selector, until=False):
                break
            if app.wait_element(
                    {"title_re": "Конфликт блокировок.*", "class_name": "", "control_type": "Pane",
                     "visible_only": True,
                     "enabled_only": True, "found_index": 0}, timeout=3):
                logger.info("Конфлик блокировок")
                if try_count < 0:
                    self.error_message = "Конфикт блокировок"
                    app.quit()
                    raise BusinessException("Не присвоен номер Конфликт блокировок", "Исходщие")
        app.quit()
        return True

    # ! переписан
    def get_outcomming_1c(self, list_of_demand):
        # ? init
        app = Odines()
        app.run()
        app.navigate("Банк и касса", "Платежное поручение исходящее", maximize_innder=True)

        # ? filter
        app.find_element({
            "title": "Установить отбор и сортировку списка...", "class_name": "", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()
        app.parent_switch({
            "title": "Отбор и сортировка", "class_name": "V8NewLocalFrameBaseWnd", "control_type": "Window",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }, resize=True)
        # * для перевода и оплаты техно
        if not app.find_element({
            "title": "Контрагент", "class_name": "", "control_type": "CheckBox",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).element.iface_toggle.CurrentToggleState:
            app.find_element({
                "title": "Контрагент", "class_name": "", "control_type": "CheckBox",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).click()
        app.find_element({
            "title": "", "class_name": "", "control_type": "Edit",
            "visible_only": True, "enabled_only": True, "found_index": 17
        }).type_keys(list_of_demand[1], app.keys.TAB, protect_first=True, click=True)
        # * для перевода
        if len(list_of_demand) == 4:
            app.find_element({
                "title": "", "class_name": "", "control_type": "Edit",
                "visible_only": True, "enabled_only": True, "found_index": 29
            }).type_keys(list_of_demand[3], app.keys.TAB, protect_first=True, click=True)
        app.find_element({
            "title": "OK", "class_name": "", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()
        if not app.wait_element({
            "title": "OK", "class_name": "", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }, until=False):
            app.close_1c_error()
            app.parent.close()
            app.quit()
            self.comments = 'Ошибка при фильтре'
            return False
        app.parent_back(1)

        # ? duplicate
        app.find_element({
            "title_re": "^.* Дата", "class_name": "", "control_type": "Custom",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).type_keys(app.keys.END, app.keys.F9, click=True, clear=False, protect_first=False)
        if app.wait_element({
            "title": "", "class_name": "", "control_type": "Document",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
        }, timeout=3):
            comment = "Документ только для чтения"
            self.comments = comment
            logger.info(comment)
            app.quit()
            return False

        # ? edit
        app.parent_switch(app.find_element({
            "title": "Платежное поручение", "class_name": "", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0}).parent(4))
        app.maximize_inner()
        app.find_element({
            "title": "Основная", "class_name": "", "control_type": "TabItem",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()
        if not app.find_element({
            "title": "Оплачено:", "class_name": "", "control_type": "CheckBox",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).element.iface_toggle.CurrentToggleState:
            app.find_element({
                "title": "Оплачено:", "class_name": "", "control_type": "CheckBox",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).click()
        keys = ['Дата выписки', 'Дата оплаты', 'Сумма платежа', 'Дополнительная информация']
        inputs = {
            e.element.element_info.element.CurrentHelpText: e
            for e in app.find_elements({
                "title": "", "class_name": "", "control_type": "Edit",
                "visible_only": True, "enabled_only": True
            }) if e.element.element_info.element.CurrentHelpText in keys
        }
        inputs['Дата выписки'].type_keys(self.operation_date, app.keys.TAB, click=True, clear=True)
        inputs['Дата оплаты'].type_keys(self.operation_date, app.keys.TAB, click=True, clear=True)
        inputs['Сумма платежа'].type_keys(self.needed_sum, app.keys.TAB, click=True, clear=True)
        comment = self.purpose_of_payment if list_of_demand[1] else ''
        inputs['Дополнительная информация'].type_keys(comment, app.keys.TAB, click=True, clear=True)

        # ? approve
        app.check_1c_error('Ошибки при заполнении')

        ok_button_selector = {"title": "ОК", "class_name": "", "control_type": "Button", "visible_only": True,
                              "enabled_only": True, "found_index": 0}
        app.find_element(ok_button_selector).click()
        # * Конфликт блокировок нужно чекнуть
        try_count = 30
        while True:
            try_count -= 1
            if app.wait_element(ok_button_selector, until=False):
                break
            if app.wait_element(
                    {"title_re": "Конфликт блокировок.*", "class_name": "", "control_type": "Pane",
                     "visible_only": True,
                     "enabled_only": True, "found_index": 0}, timeout=3):
                logger.info("Конфлик блокировок")
                if try_count < 0:
                    self.error_message = "Конфикт блокировок"
                    app.quit()
                    raise BusinessException("Не присвоен номер Конфликт блокировок", "Исходщие")
        app.quit()
        return True

    @try_except_decorator
    def process(self):

        self.start_time = time.time()
        self.retry_count = int(self.retry_count) + 1
        try:
            my_dict = {
                "Возврат оплаты за услуги операций по картам Kaspi Gold и другим картам": [
                    "Поступление оплаты по эквайрингу",
                    "KASPI BANK АО (42)",
                    1
                ],
                "Продажи c Kaspi.kz": [
                    "Поступление оплаты по эквайрингу",
                    "KASPI BANK АО (42)",
                    1
                ],
                "Продажи с Kaspi.kz": [
                    "Поступление оплаты по эквайрингу",
                    "KASPI BANK АО (42)",
                    1
                ],
                "Возврат оплаты за услуги по обработке данных": [
                    "Поступление оплаты по эквайрингу",
                    "KASPI BANK АО (42)",
                    1
                ],
                "Возврат оплаты за информационно-технологические услуги": [
                    "Возврат денежных средств поставщиком",
                    "100065350",
                    1
                ],
                "Перечисление средств за оплату парковки": [
                    "Прочее поступление безналичных денежных средств",
                    "KASPI BANK АО (42)",
                    1
                ],
                "Оплата за услуги операций по картам Kaspi Gold и другим картам": [
                    "Возврат платежа по эквайринговым операциям POS",
                    "KASPI BANK АО (42)",
                    2
                ],
                "Оплата услуги по обработке данных. В том числе НДС": [
                    "Возврат платежа по эквайринговым операциям POS",
                    "KASPI BANK АО (42)",
                    2
                ],
                "Возврат продаж с Kaspi.kz": [
                    "Возврат платежа по эквайринговым операциям POS",
                    "KASPI BANK АО (42)",
                    2
                ],
                "Комиссия за перевод": [
                    "Прочее списание безналичных денежных средств",
                    "KASPI BANK АО (42)",
                    2
                ],

                "Оплата за информационно-технологические услуги.": [
                    "Перевод на другой счет организации",
                    "100065350",
                    3
                ],
                "Перевод собственных средств на свой счет в другом Банке.": [
                    "Перевод на другой счет организации",
                    "",
                    3,
                    "KZ74722S000001851307 в АО KASPI BANK"
                ],
                "Погашение комиссии за ведение счета": [
                    "Прочее списание безналичных денежных средств",
                    "KASPI BANK АО (42)",
                    2
                ]
            }
            needed_list = []
            print(f"purp: {self.purpose_of_payment}")
            # purpose_of_payment = None
            for item in my_dict.keys():
                if str(item).lower() in str(self.purpose_of_payment).lower():
                    needed_list = my_dict.get(item)
                    break
            res = False
            if needed_list:
                if needed_list[2] == 1:
                    res = self.get_incoming_1c(needed_list)
                if needed_list[2] == 2:
                    res = self.write_off_1cc(needed_list)
                if needed_list[2] == 3:
                    res = self.get_outcomming_1c(needed_list)

            if res:
                self.status = "Success"
                self.update()
                # logger.info('Успешно создан', self.purpose_of_payment)
            else:
                # logger.info("Не удачно")
                self.status = "Retried"
                self.update()
                logger.info('Ошибка', str(self.purpose_of_payment))

        except Exception as ex:
            tb = traceback.extract_tb(ex.__traceback__)
            filename, line, func, text = tb[-1]
            msg = f'Error on line {line}: {func} {ex}'
            # logger.info(msg)
            self.status = "Retried"
            self.error_message = f"{self.error_message} {msg}"
            self.update()
            logger.info('Ошибка', str(self.purpose_of_payment))

    def update(self):
        str_now = datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        update_executor_query = f"UPDATE ROBOT.{robot_name.replace('-', '_')} SET status ='{self.status}', retry_count= {self.retry_count}, error_message ='{self.error_message}' ,comments = '{self.comments}', execution_time = '{(time.time() - self.start_time)}', finish_date = '{str_now}' WHERE id = '{self.id}' "
        print(f"update_executor_query: {update_executor_query}")
        conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
        c = conn.cursor()
        c.execute(update_executor_query)
        conn.commit()
        c.close()
        conn.close()


def notify_clients():
    # * Clients should know what is done, and what is not done
    str_today = datetime.datetime.now().strftime("%d.%m.%Y")
    select_failed_query = f"""SELECT *  FROM ROBOT.{robot_name.replace('-', '_')} where substring(date_created from 1 for 10)= '{str_today}' AND (executor_name is NULL OR executor_name = '{ip_address}') AND status ='Fail' """
    with psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass) as conn:
        with conn.cursor() as c:
            c.execute(select_failed_query)
            failed_rows = c.fetchall()
    subject = f"{robot_name_russian}"

    if len(failed_rows) > 0:
        logger.info("Есть неудачные транзакции")

        body = "Робот отработал все филиалы кроме нижеследующих \n"
        for row in failed_rows:
            body += f"{row[0]}\n"
        # logger.info(body)

    else:
        body = f"{robot_name_russian} Успешно завершился"
        logger.info("Список неудачных пуст")
    send_message_by_smtp(body=body, subject=subject, to=[to_whom, cc_whom], url=smtp_host,
                         username=smtp_author)


def download_report(report_date):
    app = Odines()
    app.run()
    app.navigate("Отчеты", "Оборотно-сальдовая ведомость по счету")
    wnd_main_report = app.find_element(
        {"title": "", "class_name": "", "control_type": "DataGrid", "visible_only": True, "enabled_only": True,
         "found_index": 0}).parent().parent().parent().parent()
    app.parent_switch(wnd_main_report)

    app.find_element(
        {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True, "enabled_only": True,
         "found_index": 2}).type_keys(report_date, app.keys.TAB, click=True, clear=True, set_focus=False)
    app.find_element(
        {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True, "enabled_only": True,
         "found_index": 3}).type_keys(report_date, app.keys.TAB, click=True, clear=True, set_focus=False)
    app.find_element(
        {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True, "enabled_only": True,
         "found_index": 0}).type_keys("1025", app.keys.ENTER, click=True, clear=True, set_focus=False)
    app.find_element(
        {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True, "enabled_only": True,
         "found_index": 1}).click()
    app.find_element(
        {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True, "enabled_only": True,
         "found_index": 1}).type_keys(app.keys.F4, click=True, clear=True, set_focus=False)

    app.parent_back(1)
    app.parent_switch(app.find_element(
        {"title": "Структурные единицы", "class_name": "V8NewLocalFrameBaseWnd", "control_type": "Window",
         "visible_only": True, "enabled_only": True, "found_index": 0, 'parent': None}))

    app.find_element({"title": "Снять флаги", "class_name": "", "control_type": "Button", "visible_only": True,
                      "enabled_only": True, "found_index": 0}).click()
    app.find_element(
        {"title": "ТОО \"Magnum Cash&Carry\" Организация", "class_name": "", "control_type": "CheckBox",
         "visible_only": True, "enabled_only": True, "found_index": 0}).click(double=True)
    # app.find_element(
    #     {"title": "ТОО \"Magnum Cash&Carry\" Организация", "class_name": "", "control_type": "CheckBox",
    #      "visible_only": True, "enabled_only": True, "found_index": 1}).click()
    # app.find_element(
    #     {"title": "ТОО \"Magnum Cash&Carry\" Организация", "class_name": "", "control_type": "CheckBox",
    #      "visible_only": True, "enabled_only": True, "found_index": 1}).type_keys(app.keys.SPACE, click=True,
    #                                                                               clear=False, set_focus=False)
    app.find_element(
        {"title": "ОК", "class_name": "", "control_type": "Button", "visible_only": True, "enabled_only": True,
         "found_index": 0}).click()

    time.sleep(5)
    app.parent_back(1)
    button_forming = app.find_element(
        {"title": "Сформировать", "class_name": "", "control_type": "Button", "visible_only": True,
         "enabled_only": True, "found_index": 0})
    button_forming.click()
    start_cursor = datetime.datetime.now()
    Cursor().wait(False)
    end_cursor = datetime.datetime.now()
    print(f"Формирование отчета закончилось за {end_cursor - start_cursor}")
    clipboard_set("")
    app.find_element(
        {"title": "", "class_name": "", "control_type": "DataGrid", "visible_only": True, "enabled_only": True,
         "found_index": 0}).type_keys('^a^c', set_focus=False)

    sTableExist = clipboard_get()
    count = 0
    while "Дебет" not in sTableExist:
        button_forming.click()
        time.sleep(40)
        clipboard_set("")
        app.find_element(
            {"title": "", "class_name": "", "control_type": "DataGrid", "visible_only": True, "enabled_only": True,
             "found_index": 0}).type_keys('^a^c', set_focus=False)
        sTableExist = clipboard_get()
        if count > 5:
            raise BusinessException("Пытались получить отчет 5 раз. Ждали по 40 секунд", "Cursor")

        count += 1

    sReportFile = temp_folder.joinpath(f"Отчет 1025_{report_date}.xlsx")

    app.find_element({"title": "Сохранить", "class_name": "", "control_type": "Button", "visible_only": True,
                      "enabled_only": True, "found_index": 0}).click()
    app.find_element(
        {"title": "Имя файла:", "class_name": "Edit", "control_type": "Edit", "visible_only": True,
         "enabled_only": True, "found_index": 0}).type_keys(sReportFile, app.keys.TAB, click=True,
                                                            set_focus=False, clear=True, protect_first=True)
    app.find_element(
        {"title": "Тип файла:", "class_name": "AppControlHost", "control_type": "ComboBox", "visible_only": True,
         "enabled_only": True, "found_index": 0}).click()
    app.find_element(
        {"title": "Лист Excel2007-... (*.xlsx)", "class_name": "", "control_type": "ListItem", "visible_only": True,
         "enabled_only": True, "found_index": 0}).click()
    app.find_element(
        {"title": "Сохранить", "class_name": "Button", "control_type": "Button", "visible_only": True,
         "enabled_only": True, "found_index": 0}).click()
    bRewrite = app.wait_element(
        {"title": "Да", "class_name": "CCPushButton", "control_type": "Button", "visible_only": True,
         "enabled_only": True, "found_index": 0}, timeout=3)
    if bRewrite:
        app.find_element(
            {"title": "Да", "class_name": "CCPushButton", "control_type": "Button", "visible_only": True,
             "enabled_only": True, "found_index": 0}).click()

    print(f"Отчет сохранен - {sReportFile}")
    app.quit()
    file_path = check_file_downloaded(Path(sReportFile))
    if not file_path:
        raise BusinessException('Файл не выгружен', f'export_1c {file_path}')
    return fix_excel_file_error(file_path)


def check_sales_report_for_new_branches(excel_file_path: str):
    # logger.info("Проверем на новые филиалы")
    sheet_name = "Выписка по продажам KaspiBank"
    # column_b = "Адрес точки продаж"
    # column_c = "Идентификатор точки"
    df = pd.read_excel(str(excel_file_path), sheet_name=sheet_name, skiprows=7)
    mapping_df = pd.read_excel(str_path_mapping_excel_file, sheet_name="Sheet1")

    mapp_column = mapping_df.columns[4]
    mapp_values = mapping_df[mapp_column].tolist()

    first_column = df.columns[1]
    sum_values = df[first_column].tolist()  # Assuming column A contains the sum values
    # row_index = 0
    # not_found_sales_point = []
    # for idx, value in enumerate(sum_values):
    #     found = False
    #     for idy, map_val in enumerate(mapp_values):
    #         if str(value) in map_val:
    #             found = True
    #             break
    #     if not found:
    #         not_found_sales_point.append(str(value))
    # list_of_non_found = ',\n'.join(set(not_found_sales_point))

    # Send email with the list of not found
    not_found_sales_point = list(set(list(r for r in sum_values if r not in mapp_values)))
    logger.info(f"Количество не найденных филиалов {len(not_found_sales_point)}")
    if not_found_sales_point:
        body = ';\n'.join(not_found_sales_point)
        subject = 'Эквайринг Каспи список не найденных фиалилов'
        try:
            send_message_by_smtp(body=body, subject=subject, to=[to_whom, cc_whom], url=smtp_host,
                                 username=smtp_author)
        except (Exception,):
            print("cannot sned email")
    return not_found_sales_point


def check_osv(report_path):
    # Read the Excel file
    df = pd.read_excel(report_path)

    # Find the row index where the 2nd column contains "Итого"
    total_row_index = df[df.iloc[:, 1] == 'Итого'].index[0]

    # Extract values from the 5th and 6th columns in the found row
    value_5th_column = df.iloc[total_row_index, 4]
    value_6th_column = df.iloc[total_row_index, 5]

    # Print the extracted values
    print(f"Value from the 5th column: {value_5th_column}")
    print(f"Value from the 6th column: {value_6th_column}")
    if float(value_5th_column) != float(value_6th_column):
        print("Суммы из ОСВ не равны")
        # TODO send email to clients
    else:
        print("Суммы ОСВ равны")


def check_sales_report(excel_file_path: str):
    # logger.info("Проверяем на пустые строки Sales Report")
    """Начинается с 7 строчки
    Проверить наличие пустых строк в колонке 1 и 2
    Если есть то отправить письмо с текстом: Обнаружены пустые строки в выписке. И далее остановить отработку данного файла.

    """
    sheet_name = "Выписка по продажам KaspiBank"
    column_b = "Адрес точки продаж"
    column_c = "Идентификатор точки"
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
    first_column = df.columns[1]
    sum_values = df[first_column].tolist()  # Assuming column A contains the sum values
    row_index = 0
    for idx, value in enumerate(sum_values):
        if column_b in str(value):
            row_index = idx
            break
    new_columns = df.iloc[row_index]
    df = df.iloc[row_index + 1:, :]
    df.columns = new_columns
    null_values_b = df[column_b].isnull().any()
    null_values_c = df[column_c].isnull().any()
    if null_values_b or null_values_c:
        print("Есть пустые строки")
        screenshot = pyautogui.screenshot()
        screenshot.save('screenshot.png')
        body = '; '.join('Каспи. Обнаружены пустые строки')
        subject = 'Эквайринг Каспи пустые строки'
        try:
            send_message_by_smtp(body=body, subject=subject, to=[to_whom, cc_whom], url=smtp_host,
                                 username=smtp_author)
        except (Exception,):
            print("cannot sned email")


@retry_n_times(3)
def upload_parking(search_date, process_date, aquir=True, real=True):
    print(upload_folder)
    app = Odines()
    app.run()

    # ? навигация
    app.navigate(
        "Сервис",
        "Внешние печатные формы, отчеты и обработки",
        "Внешние обработки"
    )
    app.find_element({
        "title_re": "^Загрузка эквайринговых операций.* Наименование", "class_name": "", "control_type": "Custom",
        "visible_only": True, "enabled_only": True, "found_index": 0
    }).click(double=True)
    app.parent_switch(app.find_element({
        "title": "Загрузка эквайринговых операций", "class_name": "", "control_type": "Text",
        "visible_only": True, "enabled_only": True, "found_index": 0
    }).parent(4))

    if aquir:
        # ? дата
        app.find_element({
            "title": "", "class_name": "", "control_type": "Edit",
            "visible_only": True, "enabled_only": True, "found_index": 1
        }).type_keys(process_date, app.keys.TAB, click=True, clear=True)

        # ? чекбокс парковка
        if not app.find_element({
            "title": "Отражение операций по услуге Парковка", "class_name": "", "control_type": "CheckBox",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).element.iface_toggle.CurrentToggleState:
            app.find_element({
                "title": "Отражение операций по услуге Парковка", "class_name": "", "control_type": "CheckBox",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).click()

        # ? файлы
        app.find_element({
            "title": "", "class_name": "", "control_type": "Edit",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).type_keys("^+{F4}", click=True, clear=True)
        app.find_element({
            "title": "Папка:", "class_name": "Edit", "control_type": "Edit",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None
        }).type_keys(upload_folder.__str__(), app.keys.ENTER, clear=True, click=True, protect_first=True)
        app.find_element({
            "title": "Выбор папки", "class_name": "Button", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None
        }).click()

        # ? подготовка
        app.find_element({
            "title": "Подготовить данные для загрузки", "class_name": "", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()
        print('start while 100% Подготовить данные для загрузки')
        while '100%' not in app.find_element({
            "title": "", "class_name": "", "control_type": "ProgressBar",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).element.iface_value.CurrentValue:
            time.sleep(1)

        flag = app.wait_element({
            "title": "", "class_name": "", "control_type": "Document",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
        }, timeout=10)
        print('Document', flag)
        if flag:
            text = app.find_element({
                "title": "", "class_name": "", "control_type": "Document",
                "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
            }).element.iface_value.CurrentValue
            if 'Загрузка прервана' in text:
                print(text)
                app.quit()
                return False

        # ? загрузка
        app.find_element({
            "title": "Выполнить загрузку", "class_name": "", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()
        if app.wait_element({
            "title": "Внимание Внимание", "class_name": "V8ConfirmationWindow", "control_type": "ToolTip",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
        }, timeout=10):
            raise Exception('ОШИБКА пустые поля при загрузке')
        print('start while Выполнить загрузку')
        while True:
            if not app.wait_element({
                "title_re": "Отчет банка по эквайрингу.*Документ отчет банка", "class_name": "",
                "control_type": "Custom",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }, timeout=300):
                app.find_element({
                    "title": "Выполнить загрузку", "class_name": "", "control_type": "Button",
                    "visible_only": True, "enabled_only": True, "found_index": 0
                }).click()
                continue
            break

    if real:
        # ? галочка парковка снять
        if app.find_element({
            "title": "Отражение операций по услуге Парковка", "class_name": "", "control_type": "CheckBox",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).element.iface_toggle.CurrentToggleState:
            app.find_element({
                "title": "Отражение операций по услуге Парковка", "class_name": "", "control_type": "CheckBox",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).click()

        # ? пропись даты реализации (дата папки)
        app.find_element({
            "title": "", "class_name": "", "control_type": "Edit",
            "visible_only": True, "enabled_only": True, "found_index": 1
        }).type_keys(search_date, app.keys.TAB, protect_first=True, click=True, clear=True)

        # ? галочка реализация
        if not app.find_element({
            "title": "Отражение операций по реализации Услуг", "class_name": "", "control_type": "CheckBox",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).element.iface_toggle.CurrentToggleState:
            app.find_element({
                "title": "Отражение операций по реализации Услуг", "class_name": "", "control_type": "CheckBox",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).click()

        # ? выполнение загрузки
        app.find_element({
            "title": "Выполнить загрузку", "class_name": "", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()
        if app.wait_element({
            "title": "Внимание Внимание", "class_name": "V8ConfirmationWindow", "control_type": "ToolTip",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
        }, timeout=10):
            raise Exception('ОШИБКА пустые поля при загрузке')
        print('start while Выполнить загрузку')
        while True:
            if app.wait_element({
                "title": "", "class_name": "", "control_type": "Document",
                "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
            }, timeout=10):
                if 'конфликт блокировок' in app.find_element({
                    "title": "", "class_name": "", "control_type": "Document",
                    "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
                }).element.iface_value.CurrentValue:
                    print('конфликт блокировок')
                    app.find_element({
                        "title": "", "class_name": "", "control_type": "Document",
                        "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
                    }).type_keys(app.keys.ESCAPE, click=True)
            if not app.wait_element({
                "title_re": "Реализация ТМЗ и услуг .* Документ отчет банка", "class_name": "",
                "control_type": "Custom",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }, timeout=300):
                app.find_element({
                    "title": "Выполнить загрузку", "class_name": "", "control_type": "Button",
                    "visible_only": True, "enabled_only": True, "found_index": 0
                }).click()
                continue
            break
    app.quit()
    return True


def split_branches(src_file, dst_dir):
    print('!!!', src_file)
    print('!!!', dst_dir)
    from pathlib import Path
    from openpyxl.reader.excel import load_workbook
    from openpyxl.workbook import Workbook

    src_file = Path(src_file)
    dst_dir = Path(dst_dir)
    wb = load_workbook(src_file.__str__())
    ws = wb.active
    vs = list(ws.values)
    start_row = [n for n, r in enumerate(vs) if r[1] == 'Адрес точки продаж'][0]

    branches = dict()
    for n, row in enumerate(vs):
        if n <= start_row:
            continue
        if not branches.get(row[1]):
            branches[row[1]] = [row]
        else:
            branches[row[1]].append(row)

    iter_data = list(branches.keys())
    iter_len = len(iter_data)
    for n, branch in enumerate(iter_data):
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = ws.title
        [new_ws.append(vs[n]) for n in range(start_row + 1)]
        [new_ws.append(r) for r in branches[branch]]
        new_wb.save(dst_dir.joinpath(f'{src_file.stem}_{n}{src_file.suffix}'))
        new_wb.close()
        print(n + 1, 'from', iter_len)
    wb.close()
    del wb
    del ws
    del vs
    del branches
    del iter_data


def prepare_upload_folder_for_one(file_full_path, split=False):
    # logger.info(f"Загружаем файл парковки: {file_full_path}")
    for filename in os.listdir(upload_folder):
        file_path = os.path.join(upload_folder, filename)

        # Check if the current item is a file
        if os.path.isfile(file_path):
            # Delete the file
            kill_process_list()
            os.unlink(file_path)
            print(f"Deleted file: {filename}")

    # logger.info("Копирование файла в папку для загрузки")
    filename = Path(file_full_path).name
    dst_file_path = upload_folder.joinpath(filename)
    try:
        if not split:
            shutil.copy(file_full_path, dst_file_path)
        else:
            split_branches(file_full_path, upload_folder)
            # Popen(['split_branches.exe', file_full_path.__str__(), upload_folder.__str__()]).wait()
        # logger.info("Файл был скопирован в папку для загрузки")
        return True
    except Exception as ex:
        traceback.print_exc()
        logger.info(f"Ошибка при копировании файла из {file_full_path} в папку {dst_file_path} {ex}")
        return False


def update_status_parking(status, retry_count, parking_id, error_message=None, comments=None):
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    c = conn.cursor()
    str_now = datetime.datetime.now().strftime("%d.%m.%Y")
    update_query = f"UPDATE ROBOT.{robot_name.replace('-', '_')}_parking SET status ='{status}', retry_count= {retry_count}, error_message ='{error_message}' ,comments = '{comments}', finish_date = '{str_now}' WHERE id = '{parking_id}' "
    c.execute(update_query)
    conn.commit()
    c.close()
    conn.close()


def upload_parking_process(aquir=True, real=True):
    while True:
        select_one_query = f"""SELECT * FROM ROBOT.{robot_name.replace('-', '_')}_parking where (executor_name is NULL OR executor_name = '{ip_address}')
                 AND status IN ('New','Retried')  ORDER BY RANDOM();"""

        conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
        c = conn.cursor()
        c.execute(select_one_query)
        row = c.fetchone()
        if row:
            # * If the retry_count exceeded  stop and set it to fail
            retry_count = int(row[2]) + 1
            if int(row[2]) > transaction_retry_count:
                update_executor_query = f"UPDATE ROBOT.{robot_name.replace('-', '_')}_parking SET executor_name ='{ip_address}', status ='Fail' WHERE id = '{row[0]}' "
                c.execute(update_executor_query)
                conn.commit()
                c.close()
                conn.close()
                continue

            else:
                update_executor_query = f"UPDATE ROBOT.{robot_name.replace('-', '_')}_parking SET executor_name ='{ip_address}' WHERE id = '{row[0]}'"
            c.execute(update_executor_query)
            conn.commit()
            c.close()
            conn.close()

            upload_folder_prepared = prepare_upload_folder_for_one(row[-1])
            if upload_folder_prepared:

                # * MAIN PARKING 1c FUNCTION

                # logger.info('Загрузка', row[10], row[9])
                result = upload_parking(row[10], row[9], aquir=aquir, real=real)

                if not result:
                    status = 'Fail'
                    logger.info('Ошибка при загрузке', row[10], row[9])
                else:
                    status = 'Success'
                    # logger.info('Успешно загружено', row[10], row[9])
                update_status_parking(status, retry_count, parking_id=row[0])

        else:
            # logger.info("Закончили загрузки парковок")
            break


def get_net_sum(file_path: str):
    """
    Используется выгруженная таблица с 1с после загрузки ТК
    """
    try:
        file_path = fix_excel_file_error(file_path)
        df = pd.read_excel(file_path, sheet_name="TDSheet")
        if len(df) == 0:
            logger.info("Sales report не загрузился в 1с, так как эксель пустлй")
            return False
        row_index: int = df['Сумма покупки'].iloc[-1:].index[0]
        purchase_sum = df.loc[row_index, 'Сумма покупки']
        return_sum = df.loc[row_index, 'Сумма возврата']
        net_sum = float(purchase_sum) - float(return_sum)
        print(f"Сумма нетто со 1с {net_sum}")
        return net_sum
    except Exception as ex:
        print(f"При получение суммы нетто с 1с, возникла ошибка {str(ex)}")
        return None


@retry_n_times(2)
def upload_sales_report_1c(process_date):
    # logger.info("Загрузка 1с Sales report")
    app = Odines()
    app.run()
    app.navigate("Сервис", "Внешние печатные формы, отчеты и обработки", "Внешние обработки")
    app.find_element(
        {"title_re": "^Загрузка эквайринговых операций.* Наименование", "class_name": "", "control_type": "Custom",
         "visible_only": True, "enabled_only": True, "found_index": 0}).click(double=True)

    acq_window_selector = {"title": "Загрузка эквайринговых операций", "class_name": "", "control_type": "Text",
                           "visible_only": True, "enabled_only": True, "found_index": 0}
    acq_window = app.wait_element(acq_window_selector, timeout=10)
    if not acq_window:
        msg = "Окно Обработка Загрузка эквайринговых операций не открылось"
        print(msg)
        raise Exception(msg)
    acq_window_element = app.find_element(acq_window_selector)
    # * Taking the 4th parent, which means the whole window
    acq_window_element = acq_window_element.parent(4)
    acq_window_element.draw_outline()
    time.sleep(3)
    app.parent_switch(acq_window_element, maximize=True)

    date_field = {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True, "enabled_only": True,
                  "found_index": 1}

    choose_catalog_field = {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True,
                            "enabled_only": True, "found_index": 0}

    app.find_element(date_field).click()
    app.find_element(date_field).type_keys(process_date, app.keys.TAB, protect_first=True, click=True, clear=True)

    app.find_element(choose_catalog_field).click()
    app.find_element(choose_catalog_field).type_keys("^+{F4}", protect_first=False, click=True, clear=True)
    time.sleep(5)
    app.parent_back(10)
    select_field = {"title": "Папка:", "class_name": "Edit", "control_type": "Edit", "visible_only": True,
                    "enabled_only": True, "found_index": 0}
    app.find_element(select_field).type_keys(str(upload_folder), app.keys.ENTER, clear=True, click=True)
    app.find_element({"title": "Выбор папки", "class_name": "Button", "control_type": "Button", "visible_only": True,
                      "enabled_only": True, "found_index": 0}).click()

    prepare_button = {"title": "Подготовить данные для загрузки", "class_name": "", "control_type": "Button",
                      "visible_only": True, "enabled_only": True, "found_index": 0}

    upload_button = {"title": "Выполнить загрузку", "class_name": "", "control_type": "Button", "visible_only": True,
                     "enabled_only": True, "found_index": 0}

    time_start = datetime.datetime.now()
    app.find_element(prepare_button).click()
    # * Ждать пока загрузит, примерно 30 min
    progress_bar_selector = {"title": "", "class_name": "", "control_type": "ProgressBar", "visible_only": True,
                             "enabled_only": True, "found_index": 0}
    count_iter = 0
    while True:
        bool_progress_bar = app.wait_element(progress_bar_selector)
        if bool_progress_bar:
            el = app.find_element(progress_bar_selector)
            value = el.element.iface_value.CurrentValue

            if value and "100%" in value:
                break

        count_iter += 1
        time.sleep(60)
        if count_iter > 360 * 3:
            print("Превышено максимальное время ожидания загрузки")
            return False
        # TODO decide how to handle this exception
    upload_time = datetime.datetime.now() - time_start
    print(f"Время загрузки {upload_time.seconds} секунд")

    # ? проверка ошибок
    if app.wait_element({
        "title": "", "class_name": "", "control_type": "Document",
        "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
    }, timeout=5):
        warning_window = app.find_element({
            "title": "", "class_name": "", "control_type": "Document",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
        })
        warning_msg = warning_window.element.iface_value.CurrentValue
        warning_window.type_keys(app.keys.ESCAPE, click=True)

        send_message_by_smtp(
            warning_msg,
            url=smtp_host,
            to=[to_whom, cc_whom],
            subject=f'Эквайринг Каспи Уведомление 1с',
            username=owa_username)
        logger.warning(f'Отправлено Уведомление 1с')
        raise Exception('Document')

    # * Нужно скачать выписку и проверить нетто сумму, совпадает ли с Х суммой
    temp_file_path = str(temp_folder.joinpath("temp.xlsx"))
    upload_window = app.find_element(
        {"title": "Предварительный просмотр", "class_name": "", "control_type": "TabItem", "visible_only": True,
         "enabled_only": True, "found_index": 0}).parent(5)
    app.parent_switch(upload_window)
    table_selector = {"title": "", "class_name": "", "control_type": "Table", "visible_only": True,
                      "enabled_only": True, "found_index": 0}

    app.find_element(table_selector).click(right=True)
    app.parent_back(1)
    app.find_element(
        {"title": "Вывести список...", "class_name": "", "control_type": "MenuItem", "visible_only": True,
         "enabled_only": True, "found_index": 0}).click()
    app.parent_switch(
        {"title": "Вывести список", "class_name": "V8NewLocalFrameBaseWnd", "control_type": "Window",
         "visible_only": True, "enabled_only": True, "found_index": 0})
    app.find_element(
        {"title": "", "class_name": "", "control_type": "Edit", "visible_only": True, "enabled_only": True,
         "found_index": 0}).type_keys(app.keys.DOWN, click=True, clear=False)
    app.find_element(
        {"title": "Табличный документ", "class_name": "", "control_type": "ListItem", "visible_only": True,
         "enabled_only": True, "found_index": 0}).click()
    app.find_element({"title": "ОК", "class_name": "", "control_type": "Button", "visible_only": True,
                      "enabled_only": True, "found_index": 0}).click()
    app.parent_back(1)

    app.find_element({"title": "", "class_name": "", "control_type": "DataGrid", "visible_only": True,
                      "enabled_only": True, "found_index": 0}).type_keys("^s", click=True, clear=False)

    app.find_element(
        {"title": "Имя файла:", "class_name": "Edit", "control_type": "Edit", "visible_only": True,
         "enabled_only": True, "found_index": 0}).type_keys(temp_file_path)
    app.find_element(
        {"title": "Тип файла:", "class_name": "AppControlHost", "control_type": "ComboBox",
         "visible_only": True, "enabled_only": True, "found_index": 0}).click()
    # * save it as xlsx
    app.find_element(
        {"title": "Лист Excel2007-... (*.xlsx)", "class_name": "", "control_type": "ListItem",
         "visible_only": True, "enabled_only": True, "found_index": 0}).click()
    # * click save button
    app.find_element(
        {"title": "Сохранить", "class_name": "Button", "control_type": "Button", "visible_only": True,
         "enabled_only": True, "found_index": 0}).click()

    doc_already_exists = app.wait_element(
        {"title": "Подтвердить сохранение в виде", "class_name": "#32770", "control_type": "Window",
         "visible_only": True, "enabled_only": True, "found_index": 0}, timeout=2)

    if doc_already_exists:
        app.find_element(
            {"title": "Да", "class_name": "CCPushButton", "control_type": "Button", "visible_only": True,
             "enabled_only": True, "found_index": 0}).click()
    doc_wnd_entire = app.find_element(
        {"title": "", "class_name": "", "control_type": "DataGrid", "visible_only": True,
         "enabled_only": True, "found_index": 0}).parent(3)

    app.parent_switch(doc_wnd_entire)
    app.find_element(
        {"title": "Закрыть", "class_name": "", "control_type": "Button", "visible_only": True,
         "enabled_only": True, "found_index": 0}).click()

    app.parent_back(1)

    # net_sum = get_net_sum(temp_file_path)

    # * После предварительной загрузки и скачивания листа нужно отправить все
    # logger.info("Clicking upload button 1c sales")
    # logger.info("!!!! НАЖАЛИ НА ВЫПОЛНИТЬ ЗАГРУЗКУ, ЖДУ КОНФЛИКТА БЛОКИРОВОК")
    app.find_element(upload_button).click()
    # input('-----')

    bottom_notification = {"title": "", "class_name": "", "control_type": "Document", "visible_only": True,
                           "enabled_only": True, "found_index": 0, "parent": app.root}

    time.sleep(3)
    while True:
        bottom_notification_appeared = app.wait_element(bottom_notification, timeout=300)
        if bottom_notification_appeared:
            text = app.find_element(bottom_notification).element.iface_value.CurrentValue
            if "конфликт блокировок" in text:
                logger.info("Конфликт блокировок")
                app.find_element(bottom_notification).type_keys(app.keys.ESCAPE, click=True)
            elif "Файл ранее загружен" in text:
                logger.info("Файл уже загружен")
                app.quit()
                return temp_file_path
        app.find_element(upload_button).click()
    # *  Тут нужно ждать сообщение и через луп каждый раз нажимать
    # popup_message_selector = {"title": "", "class_name": "", "control_type": "Document", "visible_only": True,
    #                           "enabled_only": True, "found_index": 0, "parent": app.root}
    # registry_selector = {"title_re": ".*Документ отчет банка", "class_name": "", "control_type": "Custom",
    #                      "visible_only": True, "enabled_only": True, "found_index": 3}

    # indeed_uploaded = False
    #
    # # ? грузить пока ничего не останется
    # for i in range(6*60):
    #     # * если блокировка жмем еще раз загрузить
    #     text = app.find_element(bottom_notification).element.iface_value.CurrentValue
    #     if "конфликт блокировок" in text:
    #         logger.info("Конфликт блокировок")
    #         logger.info(f"Загрузка {i + 2}")
    #         app.find_element(upload_button).click()
    #     # * если чбольше нечего грузить выход из цикла
    #     indeed_uploaded = app.wait_element(registry_selector)
    #     if not indeed_uploaded:
    #         logger.info("Успешно загрузили Sales Report")
    #         break
    #     # * прожать загрузить если не было выхода из цикла
    #     else:
    #         logger.info(f"Загрузка {i + 2}")
    #         app.find_element(upload_button).click()
    #
    # acq_window_element.close()
    #
    # if indeed_uploaded:
    #     net_sum = get_net_sum(temp_file_path)
    #     if not net_sum:
    #         myScreenshot = pyautogui.screenshot()
    #         screenshot_path = str(screenshots_folder.joinpath(
    #             f"Сделано {datetime.datetime.now().strftime('%d.%m.%Y')} за процесс день {process_date}.png"))
    #         myScreenshot.save(screenshot_path)
    #         return False
    #     print("end")
    #     # ? Осталось только одно окно Дополнительные внешние обработки
    #     app.quit()
    #
    #     return temp_file_path
    # else:
    #     logger.info("Не смогли загрузить Sales Report 10 раз")
    #     myScreenshot = pyautogui.screenshot()
    #     screenshot_path = str(screenshots_folder.joinpath(
    #         f"Сделано {datetime.datetime.now().strftime('%d.%m.%Y')} за процесс день {process_date}.png"))
    #     myScreenshot.save(screenshot_path)
    #     app.quit()
    #     # logger.info(f"Сделали скриншот")
    #     time.sleep(60 * 15)
    #     send_message_by_smtp(body='Не смогли загрузить Sales Report', subject='Ошибка Sales Report',
    #                          to=[to_whom, cc_whom], url=smtp_host, username=smtp_author,
    #                          attachments=[Path(screenshot_path)])
    #     raise BusinessException("Не смогли загрузить Sales Report", '')


def check_sales_report_vs_report_received(report_received, sales_report):
    net_sum = get_net_sum(report_received)
    if net_sum:
        try:
            df = pd.read_excel(sales_report, sheet_name="Выписка по продажам KaspiBank", skiprows=6)
            sum_of_column = df['Сумма операции (т)'].sum()
            # logger.info(f"Сумма операции (т) {sum_of_column}")
            if isinstance(sum_of_column, (int, float)):
                if net_sum != sum_of_column:
                    error_msg = f"Суммы из 1с и из Sales Report не совпадают {net_sum} {sum_of_column} соответственно"
                    logger.info(error_msg)
                    # TODO SEND EMAIL
                    # subject = robot_name_russian
                    # body = error_msg

                else:
                    logger.info(f"Суммы из 1с и из Sales Report совпадают")
            return True
        except Exception as ex:
            logger.info(f"При получение Сумма операции (т) возникла ошибка: {ex}")
            return False


def update_statuses_uploaded(rows):
    with psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass) as conn:
        with conn.cursor() as c:
            for row in rows:
                update_query = f"""Update robot.{robot_name.replace('-', '_')} set status = 'Success', uploaded='True' where id ='{row[0]}' """
                c.execute(update_query)
                conn.commit()


def perform():
    # * Reading the daily execution schedule
    wb = openpyxl.load_workbook(str_date_working_file, data_only=True)

    # Step 1: Определить рабочий день-----------------
    ws = wb['Каспи']
    operation_date = None
    today = datetime.datetime.now().strftime("%d.%m.%Y")
    # today = "12.07.2023"
    for idx, row in enumerate(ws.iter_rows(min_row=0)):
        date_m = row[0].value
        if isinstance(date_m, datetime.datetime):
            date_m = date_m.strftime("%d.%m.%Y")
        if date_m == today:
            if row[2].value == "выходной":
                wb.close()
                return
            else:
                pass
        else:
            continue
        operation_date = row[1].value
        if isinstance(operation_date, datetime.datetime):
            operation_date = operation_date.strftime("%d.%m.%Y")
        print(f"Operation_date: {operation_date}")
    wb.close()

    # Step 2: Создавать плтежное поручение
    while True:
        select_one_query = f"""SELECT * FROM ROBOT.{robot_name.replace('-', '_')} where (executor_name is NULL OR executor_name = '{ip_address}')
         AND status IN ('New','Retried')  ORDER BY RANDOM();"""

        conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
        c = conn.cursor()
        c.execute(select_one_query)
        row = c.fetchone()
        if row:
            # * If the retry_count exceeded  stop and set it to fail
            if int(row[2]) > transaction_retry_count:
                update_executor_query = f"UPDATE ROBOT.{robot_name.replace('-', '_')} SET executor_name ='{ip_address}', status ='Fail' WHERE id = '{row[0]}' "
                c.execute(update_executor_query)
                conn.commit()
                c.close()
                conn.close()
                continue
            else:
                update_executor_query = f"UPDATE ROBOT.{robot_name.replace('-', '_')} SET executor_name ='{ip_address}' WHERE id = '{row[0]}'"
            c.execute(update_executor_query)
            conn.commit()
            c.close()
            conn.close()
            tr: Transaction = Transaction(row)

            # * Main function-------------
            tr.process()
            # * Main function END ------
            tr.update()
            del tr
        else:
            break

    # Step 3: Парковка-----------------
    upload_parking_process()

    # logger.info('Запуск проверок Sales Report')

    # Step 4: Далее отрабатываем общий файл продаж по пос. терминалу
    datetime_obj = datetime.datetime.strptime(operation_date, "%d.%m.%Y")
    current_month: int = datetime_obj.month
    current_month_folder_name: str = months[current_month]
    current_year: int = datetime_obj.year
    sales_report = Path(str_sales_folder).joinpath(f"POS терминал {current_year}", current_month_folder_name)
    print(str(sales_report))
    sales_file_found = False
    for i in range(10):
        for sale_file in os.listdir(str(sales_report)):
            if operation_date in sale_file:
                sales_report = sales_report.joinpath(sale_file)
                sales_file_found = True
            print(f"sale_file: {sale_file}")
        print(sales_report)
        if not sales_file_found:
            logger.warning(f'не найден sales report от {operation_date}, ожидание 15мин')
            time.sleep(60 * 15)
        else:
            break

    # * проверка пустых
    # check_sales_report(sales_report)
    # Step 5: В сохраненном файле нужно проверить точки продаж на наличие открытия новых
    # check_sales_report_for_new_branches(sales_report)

    # Step 6: Далее загружаем Sales Report в Обработку эквайринговых операций. Указываем дату выписки, путь,
    # где расположен файл. Нажимаем Подготовить данные для загрузки
    if not sales_file_found:
        logger.info(f"Sales report за {operation_date} не найдена")
        return
    # logger.info(f'Запуск загрузки Sales Report')
    res = prepare_upload_folder_for_one(sales_report, split=True)
    if res:
        received_file = None
        for i in range(3):
            try:
                received_file = upload_sales_report_1c(process_date=operation_date)
                break
            except (Exception,):
                time.sleep(60 * 15)

        if not received_file:
            logger.info("Попробовали загрузить в 1с 3 раза. Но не получилось. Закончили работу")
            return
        # elif isinstance(received_file, bool):
        #     logger.info("Не получили файл с 1с")
        #     return
        # res = check_sales_report_vs_report_received(received_file, sales_report)
        # if res:
        #     # step 8: После выполнения загрузки нужно проверить ОСВ по счет.
        #     report_path = download_report(report_date=operation_date)
        #     check_osv(report_path)

    # notify_clients()
    logger.info(f'Загрузка и сверка завершена.\nЗавершение')


def operations(delta=0):
    # * Reading the daily execution schedule
    wb = openpyxl.load_workbook(str_date_working_file, data_only=True)

    # Step 1: Определить рабочий день-----------------
    ws = wb['Каспи']
    # operation_date = None

    today = datetime.datetime.now().date() - relativedelta(days=delta)
    today = today.strftime("%d.%m.%Y")

    for idx, row in enumerate(ws.iter_rows(min_row=0)):
        date_m = row[0].value
        if isinstance(date_m, datetime.datetime):
            date_m = date_m.strftime("%d.%m.%Y")
        if date_m == today:
            if row[2].value == "выходной":
                wb.close()
                return
            else:
                pass
        else:
            continue
        operation_date = row[1].value
        if isinstance(operation_date, datetime.datetime):
            operation_date = operation_date.strftime("%d.%m.%Y")
        print(f"Operation_date: {operation_date}")
    wb.close()

    # Step 2: Создавать плтежное поручение
    while True:
        select_one_query = f"""SELECT * FROM ROBOT.{robot_name.replace('-', '_')} where (executor_name is NULL OR executor_name = '{ip_address}')
         AND status IN ('New','Retried')  ORDER BY RANDOM();"""

        conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
        c = conn.cursor()
        c.execute(select_one_query)
        row = c.fetchone()
        if row:
            # * If the retry_count exceeded  stop and set it to fail
            if int(row[2]) > transaction_retry_count:
                update_executor_query = f"UPDATE ROBOT.{robot_name.replace('-', '_')} SET executor_name ='{ip_address}', status ='Fail' WHERE id = '{row[0]}' "
                c.execute(update_executor_query)
                conn.commit()
                c.close()
                conn.close()
                continue
            else:
                update_executor_query = f"UPDATE ROBOT.{robot_name.replace('-', '_')} SET executor_name ='{ip_address}' WHERE id = '{row[0]}'"
            c.execute(update_executor_query)
            conn.commit()
            c.close()
            conn.close()
            tr: Transaction = Transaction(row)

            # * Main function-------------
            tr.process()
            # * Main function END ------
            tr.update()
            del tr
        else:
            break


def parking(delta=0, aquir=True, real=True):
    # * Reading the daily execution schedule
    wb = openpyxl.load_workbook(str_date_working_file, data_only=True)
    # Step 1: Определить рабочий день-----------------
    ws = wb['Каспи']
    # operation_date = None

    today = datetime.datetime.now().date() - relativedelta(days=delta)
    today = today.strftime("%d.%m.%Y")

    for idx, row in enumerate(ws.iter_rows(min_row=0)):
        date_m = row[0].value
        if isinstance(date_m, datetime.datetime):
            date_m = date_m.strftime("%d.%m.%Y")
        if date_m == today:
            if row[2].value == "выходной":
                wb.close()
                return
            else:
                pass
        else:
            continue
        operation_date = row[1].value
        if isinstance(operation_date, datetime.datetime):
            operation_date = operation_date.strftime("%d.%m.%Y")
        print(f"Operation_date: {operation_date}")
    wb.close()

    upload_parking_process(aquir=aquir, real=real)


def prepare(delta=0):
    # * Reading the daily execution schedule
    wb = openpyxl.load_workbook(str_date_working_file, data_only=True)
    ws = wb['Каспи']
    operation_date = None

    today = datetime.datetime.now().date() - relativedelta(days=delta)
    today = today.strftime("%d.%m.%Y")

    # today = "04.01.2024"
    for idx, row in enumerate(ws.iter_rows(min_row=0)):
        date_m = row[0].value
        if isinstance(date_m, datetime.datetime):
            date_m = date_m.strftime("%d.%m.%Y")
        if date_m == today:
            if row[2].value == "выходной":
                wb.close()
                return
            else:
                pass
        else:
            continue
        operation_date = row[1].value
        if isinstance(operation_date, datetime.datetime):
            operation_date = operation_date.strftime("%d.%m.%Y")
        print(f"Operation_date: {operation_date}")
    wb.close()
    datetime_obj = datetime.datetime.strptime(operation_date, "%d.%m.%Y")
    current_month: int = datetime_obj.month
    current_month_folder_name: str = months[current_month]
    current_year: int = datetime_obj.year
    sales_report = Path(str_sales_folder).joinpath(f"POS терминал {current_year}", current_month_folder_name)
    print(str(sales_report))
    sales_file_found = False
    for i in range(10):
        for sale_file in os.listdir(str(sales_report)):
            if operation_date in sale_file:
                sales_report = sales_report.joinpath(sale_file)
                sales_file_found = True
            print(f"sale_file: {sale_file}")
        print(sales_report)
        if not sales_file_found:
            logger.warning(f'не найден sales report от {operation_date}, ожидание 15мин')
            time.sleep(60 * 15)
        else:
            break

    # * проверка пустых
    # check_sales_report_for_new_branches(sales_report)
    if not sales_file_found:
        logger.info(f"Sales report за {operation_date} не найдена")
        return

    # * разбивка
    res = prepare_upload_folder_for_one(sales_report, split=True)
    return res


def sales(delta=0):
    # * Reading the daily execution schedule
    wb = openpyxl.load_workbook(str_date_working_file, data_only=True)
    ws = wb['Каспи']
    operation_date = None

    today = datetime.datetime.now().date() - relativedelta(days=delta)
    today = today.strftime("%d.%m.%Y")

    for idx, row in enumerate(ws.iter_rows(min_row=0)):
        date_m = row[0].value
        if isinstance(date_m, datetime.datetime):
            date_m = date_m.strftime("%d.%m.%Y")
        if date_m == today:
            if row[2].value == "выходной":
                wb.close()
                return
            else:
                pass
        else:
            continue
        operation_date = row[1].value
        if isinstance(operation_date, datetime.datetime):
            operation_date = operation_date.strftime("%d.%m.%Y")
        print(f"Operation_date: {operation_date}")
    wb.close()
    datetime_obj = datetime.datetime.strptime(operation_date, "%d.%m.%Y")
    current_month: int = datetime_obj.month
    current_month_folder_name: str = months[current_month]
    current_year: int = datetime_obj.year
    sales_report = Path(str_sales_folder).joinpath(f"POS терминал {current_year}", current_month_folder_name)
    print(str(sales_report))
    sales_file_found = False
    for i in range(10):
        for sale_file in os.listdir(str(sales_report)):
            if operation_date in sale_file:
                sales_report = sales_report.joinpath(sale_file)
                sales_file_found = True
            print(f"sale_file: {sale_file}")
        print(sales_report)
        if not sales_file_found:
            logger.warning(f'не найден sales report от {operation_date}, ожидание 15мин')
            time.sleep(60 * 15)
        else:
            break

    received_file = None
    for i in range(3):
        try:
            received_file = upload_sales_report_1c(process_date=operation_date)
            break
        except (Exception,):
            time.sleep(60 * 15)

    if not received_file:
        logger.info("Попробовали загрузить в 1с 3 раза. Но не получилось. Закончили работу")
        return
    # # elif isinstance(received_file, bool):
    # #     logger.info("Не получили файл с 1с")
    # #     return
    # res = check_sales_report_vs_report_received(received_file, sales_report)
    # if res:
    #     # step 8: После выполнения загрузки нужно проверить ОСВ по счет.
    #     report_path = download_report(report_date=operation_date)
    #     check_osv(report_path)

    notify_clients()


if __name__ == '__main__':
    kill_process_list()
    # perform()
    parking()
    operations()
    sales()
