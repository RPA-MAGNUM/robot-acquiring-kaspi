import traceback
from contextlib import suppress
from datetime import datetime
from pathlib import Path
from shutil import copy, move
from time import sleep

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Border, Alignment, Font, Side
from openpyxl.utils import get_column_letter
from pywinauto.timings import wait_until
from sqlalchemy import create_engine, Column, String, Date, Boolean
from sqlalchemy.orm import declarative_base, sessionmaker

from config import logger, postgre_ip, postgre_port, postgre_db_name, postgre_db_username, \
    postgre_db_password, project_path, global_path, mapping_path, owa_username, smtp_host, email_to, share_path, \
    tg_logger
from core import Odines
from tools.app import MONEY_FORMAT
from tools.exceptions import BusinessException, ApplicationException
from tools.holidays import parse, names
from tools.json_rw import json_write, json_read
from tools.process import kill_process_list
from tools.smtp import smtp_send
from tools.xlsx_fix import convert, fix_excel_file_error

Base = declarative_base()


class Table(Base):
    __tablename__ = "robot_acquiring_kaspi_r"

    file_path = Column(String(512), primary_key=True)
    # id = Column(Integer, primary_key=True, autoincrement=True)
    date_created = Column(Date, default=None)

    status = Column(String(16), default=None)
    # retry_count = Column(Integer)
    comments = Column(String(256), default=None)
    fixed = Column(Boolean, default=None)
    uploaded = Column(Boolean, default=None)

    # executor_name = Column(String(16), default=None)
    # execution_time = Column(DateTime, default=None)

    date_dir = Column(String(64), default=None)
    branch_short = Column(String(64), default=None)
    branch_full = Column(String(128), default=None)
    contract = Column(String(64), default=None)
    contract_type = Column(String(64), default=None)
    # address = Column(String(256), default=None)
    transaction_sum = Column(String(64), default=None)
    comission_sum = Column(String(64), default=None)
    total_sum = Column(String(64), default=None)

    @property
    def dict(self):
        m = self.__dict__.copy()
        return m


def read_xls(reraise=True):
    while True:
        try:
            for date_dir in working_dirs:

                logger.info(f'\nзапуск {date_dir.name}')
                # ? массовый бэкап xls и конвертация в xlsx в конечной папке + заведение тасков на файл в бд -----------
                backup_dir_ = backup_dir.joinpath(
                    f'{date_dir.parents[1].name}\\{date_dir.parent.name}\\{date_dir.name}')
                # logger.info('бэкап директория для xls файлов:')
                # logger.info(backup_dir_)
                kill_process_list(['EXCEL.EXE'])
                files = [convert(f, backup_dir_, delete=True) for f in date_dir.glob('*.xls*') if 'Загрузка' not in f.name]
                # files = [convert(f, delete=True) for f in date_dir.glob('*.xls*') if 'Загрузка' not in f.name]
                logger.info(f'конвертировано {len(files)} файлов')
                reg_count = 0
                for file in files:
                    if not session.query(Table).filter_by(file_path=file.__str__()).first():
                        session.add(Table(
                            date_created=today,
                            date_dir=date_dir.name,
                            file_path=file.__str__()
                        ))
                        reg_count += 1
                if len(files):
                    session.commit()
                    logger.info(f'зарегистрировано {reg_count} задач из {date_dir.name} в бд')

                # ? правка экселей + выписка данных --------------------------------------------------------------------
                yellow_fill = PatternFill(start_color='00F5F123', fill_type='solid')
                red_fill = PatternFill(start_color='00FFCC00', fill_type='solid')

                items = [r for r in session.query(Table).filter_by(status=None, fixed=None, uploaded=None).all() if
                         r.file_path]
                for item in items:
                    path = Path(str(item.file_path))
                    wb = load_workbook(path.__str__())
                    ws = wb.active
                    values = list(ws.values)
                    # * шапка
                    name = 'Коммерсант:'
                    name_row, name_col, name_list = [(r, v.index(name), v) for r, v in enumerate(values) if name in v][
                        0]

                    # ! ЗНАЧЕНИЯ
                    contract = name_list[name_col + 1]
                    mapping_data = [
                        r for r in list(load_workbook(mapping_path).active.values)
                        if contract.strip().replace('/', '').replace('_', '').replace('-', '')
                        in str(r[2]).strip().replace('/', '').replace('_', '').replace('-', '')
                    ]
                    if not mapping_data:
                        item.contract = contract
                        item.comments = 'Контракт не найден в маппинге'
                        logger.warning(f'ОШИБКА {path.parent.name} - {path.name}\n{contract} - {item.comments}')
                        tg_logger.warning(f'ОШИБКА {path.parent.name} - {path.name}\n{contract} - {item.comments}')
                        session.commit()
                        wb.close()
                        continue
                    else:
                        _, branch_short, _, contract_type, branch_full = mapping_data[0]
                        if contract_type.lower() not in [v.lower() for v in mks]:
                            item.contract = contract
                            item.contract_type = contract_type
                            item.status = 'Skip'
                            item.comments = 'Тип контракта не подходит'
                            logger.warning(f'ПРОПУЩЕНО {path.parent.name} - {path.name}\n{contract_type} - {item.comments}')
                            tg_logger.warning(f'ПРОПУЩЕНО {path.parent.name} - {path.name}\n{contract_type} - {item.comments}')
                            session.commit()
                            wb.close()
                            continue

                    # * таблица
                    ref = 'Ret Ref Number'
                    ref_row, ref_col, ref_list = [(r, v.index(ref), v) for r, v in enumerate(values) if ref in v][0]

                    # * итого
                    itogo_row, itogo_list = [(r, v) for r, v in enumerate(values) if 'ИТОГОВЫЕ ДАННЫЕ' in v][0]
                    transaction_col = itogo_list.index('Сумма транзакций')
                    comission_col = itogo_list.index('Сумма комиссий')
                    total_col = itogo_list.index('Всего зачислено')
                    totalkz_row, totalkz_list = [(r, v) for r, v in enumerate(values) if 'Total  KZT' in v][0]

                    # ! ЗНАЧЕНИЯ
                    transaction_sum = totalkz_list[transaction_col]
                    comission_sum = totalkz_list[comission_col]
                    total_sum = totalkz_list[total_col]

                    # * Дебиторская задолженность KZT
                    debit_row, debit_list = [(r, v) for r, v in enumerate(values) if 'Дебиторская задолженность  KZT' in v][0]
                    debit_row = [debit_list[transaction_col], debit_list[comission_col], debit_list[total_col]]
                    if any(debit_row):
                        smtp_send(
                            'Обнаружена Дебиторская задолженность',
                            str(item.contract),
                            str(item.file_path),
                            str(debit_row),
                            url=smtp_host,
                            to=email_to,
                            subject='Эквайринг народный задолженность',
                            username=owa_username)
                        logger.warning('Обнаружена Дебиторская задолженность')
                        tg_logger.warning('Обнаружена Дебиторская задолженность')

                    # * выписать аномальные все
                    brokens = dict()
                    for n, row in enumerate(values):
                        if n <= ref_row:
                            continue
                        if n >= itogo_row - 1:
                            break
                        if row[ref_col] in [None, '', '\xa0']:
                            key = f'{row[ref_col - 5]}{row[ref_col - 3]}{row[ref_col - 2]}'
                            if key not in brokens:
                                brokens[key] = dict()
                            brokens[key][n] = row

                    for row_key in brokens:
                        if len(brokens[row_key].keys()) != 4:
                            continue

                        # ! покраска нормальных
                        broken = brokens[row_key]
                        correct = dict()
                        if len(broken.keys()):
                            logger.warning(f'ПРАВКА {path.parent.name} - {path.name}')
                        for n in broken:
                            if broken[n][ref_col + 1] >= 0 >= broken[n][ref_col + 2]:
                                ws.cell(n + 1, ref_col + 2).fill = yellow_fill
                                ws.cell(n + 1, ref_col + 3).fill = yellow_fill
                                ws.cell(n + 1, ref_col + 4).fill = yellow_fill
                            else:
                                correct[n] = broken[n]

                        if len(correct.values()) in [1, 3]:
                            item.comments = 'данные уже были откорректированы ранее, либо несоответствует паттерну, ' \
                                            'проверьте вручную в папке и бд'
                            logger.warning(f'ОШИБКА {path.parent.name} - {path.name}\n{item.comments}')
                            tg_logger.warning(f'ОШИБКА {path.parent.name} - {path.name}\n{item.comments}')
                            session.commit()
                            wb.close()
                            continue

                        # ! корректировка и покраска ненормальных
                        corrected = None
                        # чтобы drm обнулился, rdm заполнился
                        sorted_correct = sorted(correct, key=lambda k: correct[k][ref_col - 4])
                        for n in sorted_correct:
                            ws.cell(n + 1, ref_col + 2).fill = red_fill
                            ws.cell(n + 1, ref_col + 3).fill = red_fill
                            ws.cell(n + 1, ref_col + 4).fill = red_fill
                            if not corrected:
                                ws.cell(n + 1, ref_col + 2).value = 0
                                ws.cell(n + 1, ref_col + 3).value = 0
                                ws.cell(n + 1, ref_col + 4).value = 0
                                corrected = n
                            else:
                                ws.cell(n + 1, ref_col + 2).value += broken[corrected][ref_col + 1]
                                ws.cell(n + 1, ref_col + 3).value += broken[corrected][ref_col + 2]
                                ws.cell(n + 1, ref_col + 4).value += broken[corrected][ref_col + 3]
                                corrected = None

                    item.fixed = True
                    item.comments = None
                    item.contract = str(contract)
                    item.contract_type = str(contract_type)
                    item.branch_short = str(branch_short)
                    item.branch_full = str(branch_full)
                    item.transaction_sum = str(transaction_sum)
                    item.comission_sum = str(comission_sum)
                    item.total_sum = str(total_sum)
                    session.commit()
                    wb.save(path.__str__())
                    wb.close()
                    logger.warning(f'УСПЕШНО {path.parent.name} - {path.name}')

            # ? проверка записей
            items = session.query(Table).filter_by(status=None, fixed=None, uploaded=None).all()
            if items:
                total = [Path(str(i.file_path)).name for i in items if not i.fixed]
                err = f'{"ОШИБКА" if reraise else "ПРОПУСК"} необработано {len(total)}'
                logger.warning(err)
                tg_logger.warning(err)
                if reraise:
                    raise BusinessException(err.split('\n')[0], 'pocess')
            break

        except BusinessException:
            mapping_rows = set([
                str(r.contract)
                for r in session.query(Table).filter_by(
                    status=None, fixed=None, uploaded=None, comments='Контракт не найден в маппинге'
                ).all()
                if r.comments
            ])
            error_rows = set([
                str(r.contract)
                for r in session.query(Table).filter_by(
                    status=None, fixed=None, uploaded=None,
                    comments='данные уже были откорректированы ранее, либо несоответствует паттерну, '
                             'проверьте вручную в папке и бд'
                ).all()
                if r.comments
            ])
            smtp_send(
                'Обнаружены ошибки, робот ожидает правок 30мин.',
                'Не найдены контракты в маппинге:',
                '   ' + '\n   '.join(list(mapping_rows)),
                'Проблемы при корректировке:' if list(error_rows) else '',
                '   ' + '\n   '.join(list(mapping_rows)) if list(error_rows) else '',

                url=smtp_host,
                to=email_to,
                subject='Эквайринг народный ошибки',
                username=owa_username)
            logger.warning('выслано письмо, ждем 30мин до повторной попытки')
            tg_logger.warning('выслано письмо, ждем 30мин до повторной попытки')
            sleep(60 * 30)

        except (Exception,):
            traceback.print_exc()


def upload_tk():
    for date_dir in working_dirs:
        # ? отбор по папке с датой, тип тк, без статуса, без загрузки, с фиксом
        items = session.query(Table).filter_by(
            status=None, fixed=True, uploaded=None, date_dir=date_dir.name, contract_type='ТК'
        ).all()
        if not items:
            logger.info(f'ПРОПУЩЕНО нет файлов для загрузки {date_dir.name}')
            continue

        # ? чистка темп папки
        [p.unlink() for p in temp_dir_path.glob('*')]

        # ? копирование файлов в темповую для загрузки в 1с
        for p in items:
            copy(Path(str(p.file_path)), temp_dir_path.joinpath(Path(str(p.file_path)).name))

        # ? запуск 1с, открытие отчета
        app = Odines()
        app.run()
        app.navigate('Сервис', 'Внешние печатные формы, отчеты и обработки', 'Внешние обработки')
        app.find_element({
            "title_re": "^Загрузка эквайринговых операций.* Наименование", "class_name": "", "control_type": "Custom",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click(double=True)

        # ? пропись даты (дата папки + 1 день)
        app.parent_switch(app.find_element({
            "title": "Выполнить загрузку", "class_name": "", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).parent(6))
        date = datetime.strptime(f'{date_dir.parents[1].name}.{date_dir.name}', '%Y.%d.%m') + relativedelta(days=1)
        app.find_element({
            "title": "", "class_name": "", "control_type": "Edit",
            "visible_only": True, "enabled_only": True, "found_index": 1
        }).type_keys(date.strftime('%d.%m.%Y'), app.keys.TAB, protect_first=True, click=True, clear=True)

        # ? выбор темповой папки для загрузки
        app.find_element({
            "title": "", "class_name": "", "control_type": "Edit",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).type_keys('^+{F4}', click=True)
        app.parent_switch({
            "title": "Выбор каталога с файлами для загрузки", "class_name": "#32770", "control_type": "Window",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
        })
        app.find_element({
            "title": "Папка:", "class_name": "Edit", "control_type": "Edit",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).type_keys(str(temp_dir_path), protect_first=True, click=True)
        app.find_element({
            "title": "Выбор папки", "class_name": "Button", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()
        app.parent_back(1)
        if not app.wait_element({
            "title": "Выбор каталога с файлами для загрузки", "class_name": "#32770", "control_type": "Window",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
        }, until=False):
            raise Exception('Окно выбора папки не закрылось')

        # ? подготовка данных
        app.find_element({
            "title": "Подготовить данные для загрузки", "class_name": "", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()

        # * цикл тело для wait_until
        def wrapper():
            # * проверка конфликта блокировок, заново нажать подготовить если есть блокировка
            try:
                app.check_1c_error('check')
            except ApplicationException as e:
                if 'Конфликт блокировок' in str(e):
                    app.close_1c_error()
                    app.find_element({
                        "title": "Подготовить данные для загрузки", "class_name": "", "control_type": "Button",
                        "visible_only": True, "enabled_only": True, "found_index": 0
                    }).click()
                    return False
                else:
                    raise e
            # * проверка 100% загрузки, выход из цикла если готово
            if app.wait_element({
                "title": "", "class_name": "", "control_type": "ProgressBar",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }):
                return '100%' in app.find_element({
                    "title": "", "class_name": "", "control_type": "ProgressBar",
                    "visible_only": True, "enabled_only": True, "found_index": 0
                }).element.iface_value.CurrentValue
            else:
                return False
        # * цикл итератор
        wait_until(3600 * 3, 1, wrapper)

        # ? выгрузка файла для сверки итоговых сумм
        sleep(30)
        app.find_element({
            "title": "Предварительный просмотр", "class_name": "", "control_type": "TabItem",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()
        # sleep(30)
        app.find_element({
            "title": "", "class_name": "", "control_type": "Table",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click(right=True)
        app.find_element({
            "title": "Вывести список...", "class_name": "", "control_type": "MenuItem",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
        }).click()
        app.parent_switch({
            "title": "Вывести список", "class_name": "V8NewLocalFrameBaseWnd", "control_type": "Window",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None
        }, resize=True)
        app.find_element({
            "title": "", "class_name": "", "control_type": "Edit",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).type_keys(app.keys.DOWN, click=True)
        app.find_element({
            "title": "Табличный документ", "class_name": "", "control_type": "ListItem",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()
        app.find_element({
            "title": "ОК", "class_name": "", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()
        app.parent_back(1)
        app.find_element({
            "title": "", "class_name": "", "control_type": "DataGrid",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
        }).type_keys('^s', click=True)
        app.parent_switch({
            "title": "Сохранение", "class_name": "#32770", "control_type": "Window",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None
        })
        app.find_element({
            "title": "Тип файла:", "class_name": "AppControlHost", "control_type": "ComboBox",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).type_keys(app.keys.DOWN, click=True)
        app.find_element({
            "title": "Лист Excel2007-... (*.xlsx)", "class_name": "", "control_type": "ListItem",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()
        report_path = date_dir.joinpath(f'Загрузка {datetime.now().strftime("%Y-%m-%d %H-%M-%S")}.xlsx')
        app.find_element({
            "title": "Имя файла:", "class_name": "Edit", "control_type": "Edit",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).type_keys(report_path.__str__(), protect_first=True, click=True, clear=True)
        app.find_element({
            "title": "Сохранить", "class_name": "Button", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()
        if not app.wait_element({
            "title": "Сохранение", "class_name": "#32770", "control_type": "Window",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None
        }, until=False):
            raise Exception('Окно сохранения не закрылось')
        app.find_element({
            "title": "", "class_name": "", "control_type": "DataGrid",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
        }).type_keys(app.keys.ESCAPE, click=True, clear=True)
        if app.wait_element({
            "title": "Да", "class_name": "", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None
        }):
            app.find_element({
                "title": "Нет", "class_name": "", "control_type": "Button",
                "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None
            }).click()
        app.parent_back(1)

        # ? выполнение загрузки
        app.find_element({
            "title": "Выполнить загрузку", "class_name": "", "control_type": "Button",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }).click()
        if app.wait_element({
            "title": "Внимание Внимание", "class_name": "V8ConfirmationWindow", "control_type": "ToolTip",
            "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
        }, timeout=10):
            raise Exception('ОШИБКА пустые поля при загрузке')
        app.wait_element({
            "title_re": "^.* Документ отчет банка", "class_name": "", "control_type": "Custom",
            "visible_only": True, "enabled_only": True, "found_index": 0
        }, timeout=3600)

        # ? обход конфликта блокировок
        while True:
            if app.wait_element({
                "title": "", "class_name": "", "control_type": "Document",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }):
                text = app.find_element({
                    "title": "", "class_name": "", "control_type": "Document",
                    "visible_only": True, "enabled_only": True, "found_index": 0
                }).element.iface_value.CurrentValue
                if "конфликт блокировок" in text:
                    logger.warning("Конфликт блокировок")
                    tg_logger.warning("Конфликт блокировок")
                    app.find_element({
                        "title": "Выполнить загрузку", "class_name": "", "control_type": "Button",
                        "visible_only": True, "enabled_only": True, "found_index": 0
                    }).click()
                    continue
            break

        # ? закрытие и выход
        sleep(30)
        app.parent.close()
        app.quit()

        # ? перемещение в выполненные и пропись статуса в базе
        done_path = date_dir.joinpath('загружено')
        done_path.mkdir(exist_ok=True)
        for i in items:
            src = Path(str(i.file_path))
            dst = done_path.joinpath(src.name)
            with suppress(Exception):
                move(src, dst)
            i.status = 'Success'
            i.uploaded = True
            session.commit()
        logger.info(f'УСПЕШНО загружено {len(items)}')


def upload_parking():
    for date_dir in working_dirs:
        # ? отбор по папке с датой, тип паркинг, без статуса, без загрузки, с фиксом
        items = session.query(Table).filter_by(
            status=None, fixed=True, uploaded=None, date_dir=date_dir.name, contract_type='Паркинг'
        ).all()
        if not items:
            logger.info(f'ПРОПУЩЕНО нет файлов для загрузки паркинг {date_dir.name}')
            continue

        # ? перебор каждой парковки по отдельности
        for item in items:
            # ? чистка темп папки
            [p.unlink() for p in temp_dir_path.glob('*')]

            # ? копирование файла в темповую для загрузки в 1с
            copy(Path(str(item.file_path)), temp_dir_path.joinpath(Path(str(item.file_path)).name))

            # ? запуск 1с, открытие отчета
            app = Odines()
            app.run()
            app.navigate('Сервис', 'Внешние печатные формы, отчеты и обработки', 'Внешние обработки')
            app.find_element({
                "title_re": "^Загрузка эквайринговых операций.* Наименование", "class_name": "", "control_type": "Custom",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).click(double=True)
            app.parent_switch(app.find_element({
                "title": "Выполнить загрузку", "class_name": "", "control_type": "Button",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).parent(6))

            # ? пропись даты парковки (дата папки + 1 день)
            date = datetime.strptime(f'{date_dir.parents[1].name}.{date_dir.name}', '%Y.%d.%m') + relativedelta(days=1)
            app.find_element({
                "title": "", "class_name": "", "control_type": "Edit",
                "visible_only": True, "enabled_only": True, "found_index": 1
            }).type_keys(date.strftime('%d.%m.%Y'), app.keys.TAB, protect_first=True, click=True, clear=True)

            # ? галочка паркинг
            if not app.find_element({
                "title": "Отражение операций по услуге Парковка", "class_name": "", "control_type": "CheckBox",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).element.iface_toggle.CurrentToggleState:
                app.find_element({
                    "title": "Отражение операций по услуге Парковка", "class_name": "", "control_type": "CheckBox",
                    "visible_only": True, "enabled_only": True, "found_index": 0
                }).click()

            # ? выбор темповой папки для загрузки
            app.find_element({
                "title": "", "class_name": "", "control_type": "Edit",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).type_keys('^+{F4}', click=True)
            app.parent_switch({
                "title": "Выбор каталога с файлами для загрузки", "class_name": "#32770", "control_type": "Window",
                "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
            })
            app.find_element({
                "title": "Папка:", "class_name": "Edit", "control_type": "Edit",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).type_keys(str(temp_dir_path), protect_first=True, click=True)
            app.find_element({
                "title": "Выбор папки", "class_name": "Button", "control_type": "Button",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).click()
            app.parent_back(1)
            if not app.wait_element({
                "title": "Выбор каталога с файлами для загрузки", "class_name": "#32770", "control_type": "Window",
                "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
            }, until=False):
                raise Exception('Окно выбора папки не закрылось')

            # ? подготовка данных
            app.find_element({
                "title": "Подготовить данные для загрузки", "class_name": "", "control_type": "Button",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).click()

            # * цикл тело для wait_until
            def wrapper():
                # * проверка конфликта блокировок, заново нажать подготовить если есть блокировка
                try:
                    app.check_1c_error('check')
                except ApplicationException as e:
                    if 'Конфликт блокировок' in str(e):
                        app.close_1c_error()
                        app.find_element({
                            "title": "Подготовить данные для загрузки", "class_name": "", "control_type": "Button",
                            "visible_only": True, "enabled_only": True, "found_index": 0
                        }).click()
                        return False
                    else:
                        raise e
                # * проверка 100% загрузки, выход из цикла если готово
                if app.wait_element({
                    "title": "", "class_name": "", "control_type": "ProgressBar",
                    "visible_only": True, "enabled_only": True, "found_index": 0
                }):
                    return '100%' in app.find_element({
                        "title": "", "class_name": "", "control_type": "ProgressBar",
                        "visible_only": True, "enabled_only": True, "found_index": 0
                    }).element.iface_value.CurrentValue
                else:
                    return False
            # * цикл итератор
            wait_until(1800, 1, wrapper)

            # app.find_element({
            #     "title": "Операции по услуге Парковка", "class_name": "", "control_type": "TabItem",
            #     "visible_only": True, "enabled_only": True, "found_index": 0
            # }).click()
            # if not app.wait_element({
            #     "title": "1 N", "class_name": "", "control_type": "Custom",
            #     "visible_only": True, "enabled_only": True, "found_index": 0
            # }, timeout=5):
            #     logger.warning('ОШИБКА пустые поля при прогрузке паркинг')
            #     tg_logger.warning('ОШИБКА пустые поля при прогрузке паркинг')

            # app.find_element({
            #     "title": "Предварительный просмотр", "class_name": "", "control_type": "TabItem",
            #     "visible_only": True, "enabled_only": True, "found_index": 0
            # }).click()

            # ? выполнение загрузки
            app.find_element({
                "title": "Выполнить загрузку", "class_name": "", "control_type": "Button",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).click()
            if app.wait_element({
                "title": "Внимание Внимание", "class_name": "V8ConfirmationWindow", "control_type": "ToolTip",
                "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
            }, timeout=10):
                raise Exception('ОШИБКА пустые поля при загрузке')
            while True:
                if not app.wait_element({
                    "title_re": "Отчет банка по эквайрингу.*Документ отчет банка", "class_name": "", "control_type": "Custom",
                    "visible_only": True, "enabled_only": True, "found_index": 0
                }, timeout=300):
                    app.find_element({
                        "title": "Выполнить загрузку", "class_name": "", "control_type": "Button",
                        "visible_only": True, "enabled_only": True, "found_index": 0
                    }).click()
                    continue
                break

            # ? галочка парковка снять
            if app.find_element({
                "title": "Отражение операций по услуге Парковка", "class_name": "", "control_type": "CheckBox",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).element.iface_toggle.CurrentToggleState:
                app.find_element({
                    "title": "Отражение операций по услуге Парковка", "class_name": "", "control_type": "CheckBox",
                    "visible_only": True, "enabled_only": True, "found_index": 0
                }).click()

            # ? пропись даты реализации (дата папки)
            date = datetime.strptime(f'{date_dir.parents[1].name}.{date_dir.name}', '%Y.%d.%m')
            app.find_element({
                "title": "", "class_name": "", "control_type": "Edit",
                "visible_only": True, "enabled_only": True, "found_index": 1
            }).type_keys(date.strftime('%d.%m.%Y'), app.keys.TAB, protect_first=True, click=True, clear=True)

            # ? галочка реализация
            if not app.find_element({
                "title": "Отражение операций по реализации Услуг", "class_name": "", "control_type": "CheckBox",
                "visible_only": True, "enabled_only": True, "found_index": 0
            }).element.iface_toggle.CurrentToggleState:
                app.find_element({
                    "title": "Отражение операций по реализации Услуг", "class_name": "", "control_type": "CheckBox",
                    "visible_only": True, "enabled_only": True, "found_index": 0
                }).click()

                # ? выполнение загрузки
                app.find_element({
                    "title": "Выполнить загрузку", "class_name": "", "control_type": "Button",
                    "visible_only": True, "enabled_only": True, "found_index": 0
                }).click()
                if app.wait_element({
                    "title": "Внимание Внимание", "class_name": "V8ConfirmationWindow", "control_type": "ToolTip",
                    "visible_only": True, "enabled_only": True, "found_index": 0, "parent": app.root
                }, timeout=10):
                    raise Exception('ОШИБКА пустые поля при загрузке')
                while True:
                    if not app.wait_element({
                        "title_re": "Реализация ТМЗ и услуг .* Документ отчет банка", "class_name": "", "control_type": "Custom",
                        "visible_only": True, "enabled_only": True, "found_index": 0
                    }, timeout=300):
                        app.find_element({
                            "title": "Выполнить загрузку", "class_name": "", "control_type": "Button",
                            "visible_only": True, "enabled_only": True, "found_index": 0
                        }).click()
                        continue
                    break

            # ? закрытие и выход
            app.parent.close()
            app.quit()

            # ? перемещение в выполненные и пропись статуса в базе
            done_path = date_dir.joinpath('паркинг')
            done_path.mkdir(exist_ok=True)
            src = Path(str(item.file_path))
            dst = done_path.joinpath(src.name)
            move(src, dst)
            item.status = 'Success'
            item.uploaded = True
            session.commit()
            logger.info(f'УСПЕШНО загружен паркинг {Path(str(item.file_path)).name}')


def style_range(ws, cell_range, border=None, fill=None, font=None, alignment=None):
    if not border:
        thin = Side(border_style="thin", color="000000")
        border = Border(thin, thin, thin, thin)
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    alignment = alignment or Alignment(horizontal='center')
    font = font or Font(bold=True)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        ll = row[0]
        r = row[-1]
        ll.border = ll.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def write_xls():
    report_path = share_path.joinpath(str(today.year)).joinpath(f'выписки {today.year}.xlsx')

    thin = Side(border_style="thin", color="000000")
    border = Border(thin, thin, thin, thin)
    alignment = Alignment(horizontal='center')
    font = Font(bold=True)
    yellow_fill = PatternFill(start_color='00F5F123', fill_type='solid')
    red_fill = PatternFill(start_color='00FFCC00', fill_type='solid')

    for day_ in working_days:
        sheet_name = f'{names[day_.month - 1].lower()} {today.year}'

        # * файл, лист, шапка
        wb = load_workbook(report_path.__str__(), data_only=True) if report_path.is_file() else Workbook()
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            wb.save(report_path.__str__())
        ws = wb[sheet_name]
        ws.freeze_panes = ws['D2']
        first_day = day_.replace(day=1)
        last_day = first_day + relativedelta(months=1, days=-1)
        dates = list(first_day + relativedelta(days=d) for d in range((last_day - first_day).days + 1))
        headers = ['Контракт', 'Филиал', 'Тип']
        for n, header in enumerate(headers):
            c = ws.cell(1, n + 1)
            c.font = font
            c.alignment = alignment
            c.border = border
            c.value = header
        date_offsets = dict()
        for n, date in enumerate(dates):
            offset = (n + 1) * 3
            date_offsets[date] = offset
            ws.cell(1, offset + 1).value = date
            ws.merge_cells(start_row=1, start_column=offset + 1, end_row=1, end_column=offset + 3)
        [style_range(ws, str(r)) for r in ws.merged_cells.ranges]
        wb.save(report_path.__str__())

        # * значения
        date_dir = day_.strftime('%d.%m')
        items = session.query(Table).filter_by(uploaded=True, date_dir=date_dir).all()
        totals = [0.0, 0.0, 0.0]
        for n, item in enumerate(items):
            ws_contracts = list(r[0] for r in list(ws.values)[1:] if r[0])
            if item.contract in ws_contracts:
                contract_row = ws_contracts.index(item.contract) + 2
            else:
                contract_row = len(ws_contracts) + 2
            ws.cell(contract_row, 1).value = item.contract
            ws.cell(contract_row, 2).value = item.branch_short
            ws.cell(contract_row, 3).value = item.contract_type
            offset = date_offsets[day_]
            ws.cell(contract_row, offset + 1).value = float(str(item.transaction_sum))
            ws.cell(contract_row, offset + 1).number_format = MONEY_FORMAT
            if item.contract_type == 'Паркинг':
                ws.cell(contract_row, offset + 1).fill = red_fill
            else:
                totals[0] += float(str(item.transaction_sum))
            ws.cell(contract_row, offset + 2).value = float(str(item.comission_sum))
            ws.cell(contract_row, offset + 2).number_format = MONEY_FORMAT
            if item.contract_type == 'Паркинг':
                ws.cell(contract_row, offset + 2).fill = red_fill
            else:
                totals[1] += float(str(item.comission_sum))
            ws.cell(contract_row, offset + 3).value = float(str(item.total_sum))
            ws.cell(contract_row, offset + 3).number_format = MONEY_FORMAT
            if item.contract_type == 'Паркинг':
                ws.cell(contract_row, offset + 3).fill = red_fill
            else:
                totals[2] += float(str(item.total_sum))
            if n == len(items) - 1:
                max_row = ws.max_row
                max_row = max_row if ws.cell(max_row, 1).value is None else max_row + 1
                ws.cell(max_row, offset + 1).value = totals[0]
                ws.cell(max_row, offset + 1).number_format = MONEY_FORMAT
                ws.cell(max_row, offset + 1).fill = yellow_fill
                ws.cell(max_row, offset + 2).value = totals[1]
                ws.cell(max_row, offset + 2).number_format = MONEY_FORMAT
                ws.cell(max_row, offset + 2).fill = yellow_fill
                ws.cell(max_row, offset + 3).value = totals[2]
                ws.cell(max_row, offset + 3).number_format = MONEY_FORMAT
                ws.cell(max_row, offset + 3).fill = yellow_fill

        column_widths = []
        for row in ws.values:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell)) > column_widths[i]:
                        column_widths[i] = len(str(cell))
                else:
                    column_widths += [len(str(cell))]

        for i, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = column_width + 5
        wb.save(report_path.__str__())

        upload_report = list(Path(str(items[0].file_path)).parent.glob('Загрузка*.xlsx'))[-1]
        fix_excel_file_error(upload_report)
        upload_wb = load_workbook(upload_report.__str__(), data_only=True)
        values = list(upload_wb.active.values)
        col = values[0].index('Сумма нетто')
        total_1c = float(values[-1][col])
        total_db = round(float(totals[-1]), 2)
        if total_1c != total_db:
            smtp_send(
                f'Папка {day_.strftime("%d.%m")}',
                f'Сумма в 1с: {total_1c}',
                f'Сумма в базе: {total_db}',
                'Обнаружена разница',
                url=smtp_host,
                to=email_to,
                subject=f'Эквайринг Халык {day_.strftime("%d.%m")} РАЗНИЦА',
                username=owa_username)
            logger.warning(f'Отправлено уведомление о разнице {day_.strftime("%d.%m")}')
            tg_logger.warning(f'Отправлено уведомление о разнице {day_.strftime("%d.%m")}')

    logger.info(f'Успешно заполнен {report_path.name}')


def report():
    for day in working_days:
        query = session.query(Table).filter_by(date_dir=day.strftime('%d.%m')).all()
        success_list = list(r for r in query if r.status == 'Success')
        failed_list = list(r for r in query if r.status != 'Success')
        if not failed_list:
            json_write(last_success_path, {'day': day.strftime(day_dormat)})
        smtp_send(
            f'Папка {day.strftime("%d.%m")}',
            f'Количество успешных: {len(success_list)}',
            f'Количество неудачных: {len(failed_list)}',
            'Список неудачных:' if len(failed_list) else '',
            *[r.branch_short for r in failed_list],

            url=smtp_host,
            to=email_to,
            subject=f'Эквайринг Халык {day.strftime("%d.%m")}',
            username=owa_username)
        logger.warning(f'Отправлено уведомление о завершении {day.strftime("%d.%m")}')


if __name__ == '__main__':
    today = datetime.now().date()
    day_dormat = '%Y-%m-%d'
    logger.info(f'запуск от {today}\n')
    tg_logger.info(f'запуск от {today}\n')

    # ? проверка выходного дня -----------------------------------------------------------------------------------------
    holydays_path = global_path.joinpath(f'holydays_{today.year}.json')
    if not holydays_path.is_file():
        json_write(holydays_path, [d.strftime(day_dormat) for d in parse(today.year)])
    holydays = [datetime.strptime(d, day_dormat).date() for d in json_read(holydays_path)]
    if today in holydays:
        logger.info(f'сегодня выходной, завершение от {today}')
        tg_logger.info(f'сегодня выходной, завершение от {today}')
        exit(0)

    # ? подключение к бд -----------------------------------------------------------------------------------------------
    Session = sessionmaker()
    engine_kwargs = {
        'host': postgre_ip,
        'port': postgre_port,
        'base': postgre_db_name,
        'username': postgre_db_username,
        'password': postgre_db_password,
    }
    engine = create_engine(
        'postgresql+psycopg2://{username}:{password}@{host}:{port}/{base}'.format(**engine_kwargs),
        connect_args={'options': '-csearch_path=robot'}
    )
    Base.metadata.create_all(bind=engine)
    Session.configure(bind=engine)
    session = Session()
    logger.info('подготовка бд завершена')

    # ? подготовка папок -----------------------------------------------------------------------------------------------
    temp_dir_path = Path.home().joinpath('robot-acquiring-kaspi\\__temp__')
    temp_dir_path.mkdir(exist_ok=True, parents=True)
    backup_dir = Path.home().joinpath('robot-acquiring-kaspi\\__backup__')
    temp_dir_path.mkdir(exist_ok=True, parents=True)
    logger.info('подготовка директорий завершена')
    mks = ['ТК', 'Паркинг']

    # ? взять необходимые дни ------------------------------------------------------------------------------------------
    last_success_path = project_path.parent.joinpath('last_success_date_of_docs.json')
    if not last_success_path.is_file():
        json_write(last_success_path, {'day': (today - relativedelta(days=3)).strftime(day_dormat)})
    last_success_day = datetime.strptime(json_read(last_success_path).get('day'), day_dormat).date()
    working_days = [(last_success_day + relativedelta(days=x + 1)) for x in range((today - last_success_day).days - 2)]
    working_dirs = list()
    for d in working_days:
        y_path = share_path.joinpath(f'{d.year}')
        try:
            m_path = list(
                x for x in y_path.glob(f'*{d.month}.*')
                if x.is_dir() and names[d.month - 1].lower() in x.name.lower()
            )[0]
            d_path = m_path.joinpath(d.strftime('%d.%m'))
            if not d_path.is_dir():
                msg = f'ОШИБКА папка не найдена {d_path.name}'
                logger.info(d_path, msg)
                tg_logger.info(d_path, msg)
                raise Exception(msg)
            working_dirs.append(d_path)
        except (Exception,):
            # traceback.print_exc()
            continue
    if len(working_dirs):
        logger.info('дни для обработки:')
        logger.info('\n'.join(p.__str__() for p in working_dirs))
    else:
        logger.warning('ОШИБКА не найдены папки для отработки', '\n')
        tg_logger.warning('ОШИБКА не найдены папки для отработки', '\n')

    # ? подготовка экселей на каждую дату
    read_xls(reraise=True)

    # ? загрузка ТК на каждую дату
    upload_tk()

    # ? загрузка Паркинг каждую дату
    upload_parking()

    # ? проверка бд, запись успешного дня, запись отчета
    write_xls()

    # ? сдвиг успешного дня + оповещение
    report()

    logger.info(f'завершение от {today}')
    tg_logger.info(f'завершение от {today}')
