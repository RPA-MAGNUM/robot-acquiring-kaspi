import datetime
from datetime import datetime
from pathlib import Path
from shutil import move

import requests
from bs4 import BeautifulSoup
from dateutil.relativedelta import relativedelta
from dateutil.rrule import rrule, DAILY
from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from rpamini import json_write, json_read

names = (
    'Январь', 'Февраль', 'Март',
    'Апрель', 'Май', 'Июнь',
    'Июль', 'Август', 'Сентябрь',
    'Октябрь', 'Ноябрь', 'Декабрь'
)


def parse(year: int):
    webpage = requests.get(f'https://online.zakon.kz/accountant/Calendars/Holidays/{year}', verify=False)
    soup = BeautifulSoup(webpage.content, "html.parser")
    dom = etree.HTML(str(soup))
    months = dom.xpath('//div[@class="month-col--calendar"]')

    data = dict()
    total_count = 0
    for month in months:
        month_name = month.xpath('.//h6')[0].text
        holidays = month.xpath(
            './/div[not(contains(@class, "prev-month")) and contains(@class, "calendar-day") '
            'and contains(@class, "holiday")]/span'
        )
        idx = names.index(month_name) + 1
        data[idx] = [int(r.text.strip()) for r in holidays]
        total_count += len(data[idx])

    web_holidays_count = int(dom.xpath('//div[contains(text(), "выходных дней")]/h2')[0].text)
    web_superholidays_count = int(dom.xpath('//div[contains(text(), "праздничных дней")]/h2')[0].text)
    if total_count != web_holidays_count + web_superholidays_count:
        raise Exception('Количество с портала не соответствует выгруженному количеству')

    result = list()
    for month in data:
        for day in data[month]:
            result.append(datetime(year, month, day).date())
    return result


def generate(mapping_path: Path):
    df = '%d.%m.%Y'
    title = 'Каспи'
    if mapping_path.is_file():
        wb = load_workbook(mapping_path.__str__())
        ws = wb[title]
        values = list(ws.values)
        date = datetime.strptime(values[1][0], df)
        if date.year != datetime.now().year:
            move(mapping_path, mapping_path.parent.joinpath(f'{mapping_path.stem}_{date.year}{mapping_path.suffix}'))

    if mapping_path.is_file():
        return

    path = Path('holidays.json').absolute()
    if not path.is_file():
        year = datetime.now().year
        json_write(path, [r.strftime(df) for r in [*parse(year - 1), *parse(year)]])
    holidays = [datetime.strptime(r, df) for r in json_read(path)]

    wb = Workbook()
    ws = wb.active
    ws.title = title

    ws.append([
        'Дата запуска',
        'Дата операции(в обработке Загрузка экв.операций) 1С',
        'Название ваыписки по продажам',
        'Дата операции в загрузке в 1с',
        'Парковка, галочка, дата выписки название',
        'Реал галочка, дата операции'
    ])

    date_from = datetime.now().date().replace(day=1, month=1)
    date_to = date_from + relativedelta(years=1, days=-1)
    date_range = [d for d in rrule(DAILY, dtstart=date_from - relativedelta(days=30), until=date_to)]
    date_dict = {
        d: 'holiday' if d in holidays else
        'first_after' if d - relativedelta(days=1) in holidays else
        'second_after' if d - relativedelta(days=2) in holidays else
        'normal'
        for d in date_range
    }
    for k in date_dict:
        if k.year != datetime.now().year:
            continue
        # print(k, date_range[k])
        if date_dict[k] == 'holiday':
            ws.append([
                k.strftime(df),
                'выходной',
                'выходной',
                'выходной',
                'выходной',
                'выходной',
            ])
        elif date_dict[k] == 'first_after':
            last = [k - relativedelta(days=i) for i in range(1, 30) if k - relativedelta(days=i) not in holidays][0]
            if date_dict[last] == 'first_after':
                days = [last, last - relativedelta(days=1)]
                for i in range(2, 30):
                    if last - relativedelta(days=i) in holidays:
                        days.append(last - relativedelta(days=i))
                    else:
                        break
                ws.append([
                    k.strftime(df),
                    days[0].strftime(df),
                    ';'.join([d.strftime(df) for d in reversed(days)]),
                    ';'.join([(d - relativedelta(days=1)).strftime(df) for d in reversed(days)]),
                    days[0].strftime(df),
                    ';'.join([(d - relativedelta(days=1)).strftime(df) for d in reversed(days)]),
                ])
            else:
                ws.append([
                    k.strftime(df),
                    last.strftime(df),
                    last.strftime(df),
                    (last - relativedelta(days=1)).strftime(df),
                    last.strftime(df),
                    (last - relativedelta(days=1)).strftime(df),
                ])
        elif date_dict[k] == 'second_after':
            days = [k - relativedelta(days=1)]
            for i in range(2, 30):
                if k - relativedelta(days=i) in holidays:
                    days.append(k - relativedelta(days=i))
                else:
                    break
            ws.append([
                k.strftime(df),
                days[0].strftime(df),
                ';'.join([d.strftime(df) for d in reversed(days)]),
                ';'.join([(d - relativedelta(days=1)).strftime(df) for d in reversed(days)]),
                days[0].strftime(df),
                ';'.join([(d - relativedelta(days=1)).strftime(df) for d in reversed(days)]),
            ])
        else:
            ws.append([
                k.strftime(df),
                (k - relativedelta(days=1)).strftime(df),
                (k - relativedelta(days=1)).strftime(df),
                (k - relativedelta(days=2)).strftime(df),
                (k - relativedelta(days=1)).strftime(df),
                (k - relativedelta(days=2)).strftime(df),
            ])

    for n, column in enumerate(list(ws.columns), 1):
        ws.column_dimensions[get_column_letter(n)].width = max([len(c.value) for c in column])
    wb.save(mapping_path.__str__())
    wb.close()


if __name__ == '__main__':
    generate(Path(r"\\172.16.8.87\d\.rpa\.agent\robot-acquiring-kaspi\маппинг загрузка файлов.xlsx"))
